from app.helpers.common import *
from app.settings import *
import os
from app.helpers.common import *

# sync_dir_structure(JAR_DIR)
jar_dir = JAR_DIR
# sync_dir_structure(TRANSFORMATION_OUTPUT_DIR)
output_dir = TRANSFORMATION_OUTPUT_DIR
template_dir = TRANSFORMATION_TEMPLATEFILES_DIR

class PreviewRepoService:
    def imports():
        '''
        This function returns the import structure for every python file to be created.

        '''
        lines = f"""import pandas as pd
from openpyxl import load_workbook
import functools
import builtins
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import logging
from functools import reduce
from datetime import datetime
import builtins
import paramiko
from io import StringIO
import tempfile

def save_dataframes_to_excel(dataframes, sheet_names, excel_filename):
    '''
    Save multiple pandas DataFrames into a single Excel file with multiple sheets.

    :param dataframes: List of pandas DataFrames
    :param sheet_names: List of sheet names corresponding to each DataFrame
    :param excel_filename: Name of the Excel file to create
    '''
    try:
        # Load the existing workbook if it exists
        workbook = load_workbook(filename=excel_filename)
    except FileNotFoundError:
        # If the workbook doesn't exist, create a new one
        workbook = openpyxl.Workbook()
        # Remove the default sheet created by openpyxl
        default_sheet = workbook.active
        print(default_sheet)
        workbook.remove(default_sheet)

    # Add each DataFrame to the Excel workbook as a separate sheet
    for df, sheet_name in zip(dataframes, sheet_names):
        print("Inside For Loop")
        # Create a new sheet with the given name
        worksheet = workbook.create_sheet(title=sheet_name)
        # Convert DataFrame to rows and write to Excel sheet
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

    # Save the workbook to the specified file
    print("Saving Workbook")
    workbook.save(excel_filename)

def save_dataframes_to_excel_1(dataframes, sheet_names, excel_filename):
    '''
    Save multiple PySpark DataFrames into a single Excel file with multiple sheets.

    :param dataframes: List of PySpark DataFrames
    :param sheet_names: List of sheet names corresponding to each DataFrame
    :param excel_filename: Name of the Excel file to create
    '''
    try:
        try:
            # Load the existing workbook if it exists
            workbook = load_workbook(filename=excel_filename)
        except FileNotFoundError:
            # If the workbook doesn't exist, create a new one
            workbook = openpyxl.Workbook()

        # Add each DataFrame to the Excel workbook as a separate sheet
        for df, sheet_name in zip(dataframes, sheet_names):
            # Check if the sheet already exists
            # Create a new sheet with the given name
            worksheet = workbook.create_sheet(title=sheet_name)
            # Collect the data from the DataFrame
            data = df.collect()  # Collect the data locally
            # Convert the data into rows that can be added to the Excel sheet
            # Get column names
            col_names = df.columns
            # Add column headers to the sheet
            worksheet.append(col_names)
            # Add data rows to the sheet
            for row in data:
                worksheet.append(row)

        # Save the workbook to the specified file
        workbook.save(excel_filename)
        df = pd.DataFrame({{'Output': ['Successfully executed Saving of zip files']}})
        return df
    except Exception as e:
        df = pd.DataFrame({{'Error': [f'{{str(e)}}']}})
        return df
"""
        return lines
    
    def readExcel(**kwargs):
        try:
            requiredKeys = ['sheetName', 'df', 'localFilePath', 'fileName']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for key, value in zip(requiredKeys, values) if key != 'fileName'):
                    sheetName, df, localFilePath, fileName = values
                    if fileName == '':
                        fileName = ''
                    else:
                        fileName = fileName
                        
                    lines = f"""
# Read the Excel file into a pandas DataFrame
{df} = pd.read_excel(f"{localFilePath.strip()}/{fileName.split('/')[-1]}", sheet_name="{sheetName}")

# Remove all periods from the column names
{df}.columns = {df}.columns.str.replace('.', '')

# Filter out columns that start with "Unnamed"
columns_to_keep = [col for col in {df}.columns if not col.startswith("Unnamed")]

# Create a new DataFrame with only the columns to retain
{df} = {df}[columns_to_keep]
"""
                    logger.info(lines)
                    return lines, True
                else:
                    return "filePath, fileName or sheetName not provided", False
                
            else:
                return "filePath or sheetName not provided", False
        
        except Exception as e:
            return str(e), False
        
    def readCsv(**kwargs):
        try:
            requiredKeys = ['fileName', 'delimiter', 'df', 'localFilePath']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for key, value in zip(requiredKeys, values) if key != 'fileName'):
                    fileName, delimiter, df, localFilePath = values
                    if fileName == '':
                        fileName = ''
                    else:
                        fileName = fileName
                    lines = f"""
# Read the CSV file into a pandas DataFrame
{df} = pd.read_csv(f'{localFilePath}/{fileName.split('/')[-1]}', delimiter='{delimiter}')

# Remove all periods from the column names
{df}.columns = {df}.columns.str.replace('.', '', regex=False)

# Filter out columns that start with "Unnamed"
columns_to_keep = [col for col in {df}.columns if not col.startswith("Unnamed")]

# Create a new DataFrame with only the columns to retain
{df} = {df}[columns_to_keep]
"""
                    logger.info(lines)
                    return lines, True
                else:
                    return "filePath not provided", False
            else:
                return "filePath or fileName not provided", False
        
        except Exception as e:
            return str(e), False
        
    def selectCols(dropna, **kwargs):
        try:
            newDf_found = False
            oldDf_found = False
            columnString_found = False

            for key, value in kwargs.items():
                if key == 'newDf':
                    newDf_found = True
                    newDf = value
                if key == 'oldDf':
                    oldDf_found = True
                    oldDf = value
                if key == 'columnString':
                    columnString_found = True
                    columnString = value

            if columnString_found:
                col_list = columnString.split(',')
                col_list = list(map(lambda x: x.strip(), col_list))
                rename_dict = {}
                for col in col_list:
                    if ' as ' in col:
                        col = col.split(' as ')
                        rename_dict.update({col[0]: col[1]})
                    else:
                        rename_dict.update({col.strip(): col.strip()})
                
                if (oldDf_found & newDf_found):
                    if dropna:
                        pyspark_command = f"""
{newDf} = {oldDf}[list({rename_dict}.keys())].rename(columns={rename_dict})
# Drop rows with NaN values in the selected columns
{newDf} = {newDf}.dropna()
"""
                    else:
                        pyspark_command = f"""
{newDf} = {oldDf}[list({rename_dict}.keys())].rename(columns={rename_dict})
"""
                    logger.info(pyspark_command)
                    return pyspark_command, True
                else:
                    return "Key Value pair errors", False
            else:
                return "Key Value pair errors selectCols function", False
        
        except Exception as e:
            return str(e), False
        
    def concatWs(**kwargs):
        try:
            requiredKeys = ['delimiter', 'columnString','df','newCol']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                delimiter, columnString, df, newCol = values
                columnString = columnString.split(',')
                column_list = list(map(lambda x: x.strip(), columnString))

                pyspark_command = f"""
################ command for concat_ws ###################
# Create a new column by concatenating the specified columns with the delimiter
{df}['{newCol}'] = {df}[{column_list}].apply(lambda x: '{delimiter}'.join(x.astype(str)), axis=1)
"""
                logger.info(pyspark_command)
                return pyspark_command, True
            else:
                    return "Key Value pair errors in concatWs function", False
            
        except Exception as e:
            return str(e), False
        
    def renameColumns(**kwargs):
        try:
            requiredKeys = ['df', 'renameColumnString']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    df, renameColumnString = values
                    renameColumnString = renameColumnString.split(',')
                    renameColumnString = list(map(lambda x: x.strip(), renameColumnString))
                    renamed_cols = {}
                    for renamed_string in renameColumnString:
                        renamed_string = renamed_string.split(' as ')
                        renamed_cols.update({renamed_string[0]: renamed_string[1]})
                    
                    line_item = f"""
{df} = {df}.rename(columns = {renamed_cols})
"""
                    logger.info(line_item)
                    return line_item, True
                else:
                    return "Key Value pair errors in rename column errors", False
            else:
                return "Key Value pair errors in rename column errors", False
        except Exception as e:
            return str(e), False
        
    def selectLeft(**kwargs):
        try:
            requiredKeys = ['df', 'oldColumn', 'newColumn', 'endIndex']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    df, oldColumn, newColumn, endIndex = values
                    lines = f"""
{df}['{newColumn}'] = {df}['{oldColumn}'].str[:{endIndex}]
"""
                    logger.info(lines)
                    return lines, True
                else:
                    return "Key Value pair errors select left function", False
            else:
                return "Key Value pair errors select left function", False
        
        except Exception as e:
            return str(e), False
        
    def selectRight(**kwargs):
        try:
            requiredKeys = ['df', 'oldColumn', 'newColumn', 'startIndex']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    df, oldColumn, newColumn, startIndex = values
                    lines = f"""
{df}['{newColumn}'] = {df}['{oldColumn}'].str[-{startIndex}:]
"""
                    return lines, True
                else:
                    return "Key Value pair errors select right function", False
            else:
                return "Key Value pair errors select right function", False
        
        except Exception as e:
            return str(e), False
        
    def filterDf(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'filterConditions']
            if all(key in kwargs.keys() for key in requiredKeys):
                oldDf, newDf, filterConditions = [kwargs.get(key) for key in requiredKeys]
                # Start by copying the old DataFrame to new DataFrame
                filter_conditions_list = []
                for condition in filterConditions:
                    if 'group' in condition:
                        sub_conditions = []
                        for sub_condition in condition['group']:
                            if not sub_condition['isNumber']:
                                value = f"'{sub_condition['value']}'"
                            else:
                                value = f"{sub_condition['value']}"
                        
                            if sub_condition['logicalOperator'] == 'AND':
                                logical_operator = ' & '
                            elif sub_condition['logicalOperator'] == 'OR':
                                logical_operator = ' | '
                            else:
                                logical_operator = ''

                            if (sub_condition['operator'] == 'contains') or (sub_condition['operator'] == 'like'):
                                value = f"'{sub_condition['value']}'"
                                sub_condition_line = f"""({oldDf}["{sub_condition['columnName']}"].str.contains({value}, na = False)){logical_operator}"""
                            elif sub_condition['operator'] == 'rlike':
                                value = f"'{sub_condition['value']}'"
                                sub_condition_line = f"""({oldDf}["{sub_condition['columnName']}"].str.contains(r"^{sub_condition['value']}", regex = True, na = False)){logical_operator}"""
                            elif sub_condition['operator'] == 'isNull':
                                sub_condition_line = f"""({oldDf}["{sub_condition['columnName']}"].isna(){logical_operator}"""
                            elif sub_condition['operator'] == 'isNotNull':
                                sub_condition_line = f"""({oldDf}["{sub_condition['columnName']}"].notna(){logical_operator}"""
                            else:
                                if sub_condition['columnOperation'] != '':
                                    sub_condition_line = f"""({oldDf}["{sub_condition["columnName"]}"].str.len() {sub_condition["operator"]} {value}){logical_operator}"""
                                else:
                                    sub_condition_line = f"""({oldDf}["{sub_condition["columnName"]}"] {sub_condition["operator"]} {value}){logical_operator}"""
                            
                            sub_conditions.append(sub_condition_line)

                            condition_line = '(' + ''.join(sub_conditions) + ')'

                    else:
                        if not condition['isNumber']:
                            value = f"'{condition['value']}'"
                        else:
                            value = f"{condition['value']}"

                        if (condition['operator'] == 'contains') or (condition['operator'] == 'like'):
                            value = f"'{condition['value']}'"
                            condition_line = f"""({oldDf}["{condition['columnName']}"].str.contains({value}, na = False))"""
                        elif condition['operator'] == 'rlike':
                            value = f"'{condition['value']}'"
                            condition_line = f"""({oldDf}["{condition['columnName']}"].str.contains(r"^{condition['value']}", regex = True, na = False))"""
                        elif condition['operator'] == 'isNull':
                            condition_line = f"""({oldDf}["{condition['columnName']}"].isna()"""
                        elif condition['operator'] == 'isNotNull':
                            condition_line = f"""({oldDf}["{condition['columnName']}"].notna()"""
                        else:
                            if condition['columnOperation'] != '':
                                condition_line = f"""({oldDf}["{condition["columnName"]}"].str.len() {condition["operator"]} {value})"""
                            else:
                                condition_line = f"""({oldDf}["{condition["columnName"]}"] {condition["operator"]} {value})"""
                    
                    filter_conditions_list.append(condition_line)
                
                lines = f"{newDf} = {oldDf}[" + " & ".join(filter_conditions_list) + "]"
                return lines, True
            else:
                return "Key value pair errors in filter_df", False
        except Exception as e:
            return str(e), False
        
    def filterDelete(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'filterConditions']
            if all(key in kwargs.keys() for key in requiredKeys):
                oldDf, newDf, filterConditions = [kwargs.get(key) for key in requiredKeys]
                # Start by copying the old DataFrame to new DataFrame
                filter_conditions_list = []
                for condition in filterConditions:
                    if 'group' in condition:
                        sub_conditions = []
                        for sub_condition in condition['group']:
                            if not sub_condition['isNumber']:
                                value = f"'{sub_condition['value']}'"
                            else:
                                value = f"{sub_condition['value']}"
                        
                            if sub_condition['logicalOperator'] == 'AND':
                                logical_operator = ' & '
                            elif sub_condition['logicalOperator'] == 'OR':
                                logical_operator = ' | '
                            else:
                                logical_operator = ''

                            if (sub_condition['operator'] == 'contains') or (sub_condition['operator'] == 'like'):
                                value = f"'{sub_condition['value']}'"
                                sub_condition_line = f"""({oldDf}["{sub_condition['columnName']}"].str.contains({value}, na = False)){logical_operator}"""
                            elif sub_condition['operator'] == 'rlike':
                                value = f"'{sub_condition['value']}'"
                                sub_condition_line = f"""({oldDf}["{sub_condition['columnName']}"].str.contains(r"^{sub_condition['value']}", regex = True, na = False)){logical_operator}"""
                            elif sub_condition['operator'] == 'isNull':
                                sub_condition_line = f"""({oldDf}["{sub_condition['columnName']}"].isna(){logical_operator}"""
                            elif sub_condition['operator'] == 'isNotNull':
                                sub_condition_line = f"""({oldDf}["{sub_condition['columnName']}"].notna(){logical_operator}"""
                            else:
                                if sub_condition['columnOperation'] != '':
                                    sub_condition_line = f"""({oldDf}["{sub_condition["columnName"]}"].str.len() {sub_condition["operator"]} {value}){logical_operator}"""
                                else:
                                    sub_condition_line = f"""({oldDf}["{sub_condition["columnName"]}"] {sub_condition["operator"]} {value}){logical_operator}"""
                            
                            sub_conditions.append(sub_condition_line)

                            condition_line = '~(' + ''.join(sub_conditions) + ')'

                    else:
                        if not condition['isNumber']:
                            value = f"'{condition['value']}'"
                        else:
                            value = f"{condition['value']}"

                        if (condition['operator'] == 'contains') or (condition['operator'] == 'like'):
                            value = f"'{condition['value']}'"
                            condition_line = f"""(~{oldDf}["{condition['columnName']}"].str.contains({value}, na = False))"""
                        elif condition['operator'] == 'rlike':
                            value = f"'{condition['value']}'"
                            condition_line = f"""(~{oldDf}["{condition['columnName']}"].str.contains(r"^{condition['value']}", regex = True, na = False))"""
                        elif condition['operator'] == 'isNull':
                            condition_line = f"""(~{oldDf}["{condition['columnName']}"].isna()"""
                        elif condition['operator'] == 'isNotNull':
                            condition_line = f"""(~{oldDf}["{condition['columnName']}"].notna()"""
                        else:
                            if condition['columnOperation'] != '':
                                condition_line = f"""(~{oldDf}["{condition["columnName"]}"].str.len() {condition["operator"]} {value})"""
                            else:
                                condition_line = f"""(~{oldDf}["{condition["columnName"]}"] {condition["operator"]} {value})"""
                    
                    filter_conditions_list.append(condition_line)
                
                lines = f"{newDf} = {oldDf}[" + " & ".join(filter_conditions_list) + "]"
                return lines, True
            else:
                return "Key value pair errors in filterDelete()", False
        except Exception as e:
            return f"Error in PreviewRepoService filterDelete() :: {str(e)}", False
        
    def columnTrim(**kwargs):
        try:
            requiredKeys = ['df', 'oldColumn', 'newColumn']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    df, oldColumn, newColumn = values

                    lines = f"""
# Function to clean up spacing issues
def clean_description(text):
    # Remove leading/trailing spaces and ensure only one space between words
    return re.sub(r'\s+', ' ', text).strip()

# Apply the function to the 'Description' column
{df}['{newColumn}'] = {df}['{oldColumn}'].apply(clean_description)
{df}['{newColumn}'] = {df}['{newColumn}'].str.strip()
"""

                    return lines, True
                else: 
                    return "Key Value pair empty in column trim function", False
            else:
                return "Key Value pair error in column trim function", False
        except Exception as e:
            return str(e), False
        
    def getLength(**kwargs):
        try:
            requiredKeys = ['df', 'oldColumn', 'newColumn']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    df, oldColumn, newColumn = values

                    lines = f"""{df}['{newColumn}'] = {df}['{oldColumn}'].str.len()"""

                    return lines, True
                else:
                    return "Key Value pair empty in get Length Function", False
            else:
                return "Key Value pair errors in getLength", False
        except Exception as e:
            return str(e), False
        
    def dropDuplicates(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'columnString']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    oldDf, newDf, columnString = values
                    columnString = columnString.split(',')
                    columnString_list = list(map(lambda x: x.strip(), columnString))
                    lines = f"""{newDf} = {oldDf}.drop_duplicates(subset = {columnString_list})"""

                    return lines, True
                else:
                    return "Key value pair empty in dropDuplicates function", False
            else:
                return "Key Value pair errors in drop duplicates", False

        except Exception as e:
            return str(e), False
        
    def createNewColumnWithDefaultValue(**kwargs):
        try:
            requiredKeys = ['df', 'defaultValue', 'columnName']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    df, defaultValue, columnName = values

                    lines = f"""{df}['{columnName}'] = '{defaultValue}'"""
                    return lines, True
                else:
                    return "Key Value pair empty in createNewColumnWithDefaultValue", False
            else:
                return "key value pair error in createNewColumnWithDefaultValue function", False
        except Exception as e:
            return str(e), False
        
    def copyColToNewCol(**kwargs):
        try:
            requiredKeys = ['df', 'oldColumn', 'newColumn']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    df, oldColumn, newColumn = values

                    lines = f"""{df}['{newColumn}'] = {df}['{oldColumn}']"""

                    return lines, True
                else:
                    return "Key value pairs empty in copyColToNewCol function", False
            else:
                return "Key Value pair errors in copy to new column function", False
        except Exception as e:
            return str(e), False
        
    def dropnaSubset(**kwargs):
        try:
            requiredKeys = ['df', 'columnString']
            if all(key in kwargs.keys() for key in requiredKeys):
                df = kwargs['df']
                columnString = kwargs['columnString']

                columnString = columnString.split(',')
                column_list = list(map(lambda x: x.strip(), columnString))

                lines = f"""{df} = {df}.dropna(subset = {column_list})"""
                return lines, True
            else:
                return "Key Value pair error in dropna Subset Function", False
        except Exception as e:
            return str(e), False
        
    def leadColumn(**kwargs):
        try:
            requiredKeys = ['df','oldColumn', 'newColumn', 'leadBy']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, oldColumn, newColumn, leadBy = [kwargs.get(key) for key in requiredKeys]

                lines = f"""{df}['{newColumn}'] = {df}['{oldColumn}'].shift(-{leadBy})"""
                return lines, True
            else:
                return "Key Value pair error in leadColumn Function", False
        except Exception as e:
            return str(e), False
        
    def lagColumn(**kwargs):
        try:
            requiredKeys = ['df','oldColumn', 'newColumn', 'lagBy']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, oldColumn, newColumn, lagBy = [kwargs.get(key) for key in requiredKeys]

                lines = f"""{df}['{newColumn}'] = {df}['{oldColumn}'].shift({lagBy})"""
                return lines, True
            else:
                return "Key Value pair error in lagColumn Function", False
        except Exception as e:
            return str(e), False
        
    def dateDiff(**kwargs):
        try:
            requiredKeys = ['resultDf', 'df', 'newColumn', 'oldColumn', 'lastDay']
            if all(key in kwargs.keys() for key in requiredKeys):
                resultDf, df, newColumn, oldColumn, lastDay = [kwargs.get(key) for key in requiredKeys]
                
                # Convert oldColumn to datetime if not already
                lines = f"""
{resultDf} = {df}.copy()
{resultDf}['{oldColumn}'] = pd.to_datetime({resultDf}['{oldColumn}'])
"""
                
                if lastDay:
                    # Get the last day of the current month
                    lines += f"""
current_date = pd.Timestamp(datetime.now())
last_day = current_date.replace(day=1) + pd.offsets.MonthEnd(1)
{resultDf}['{newColumn}'] = (last_day - {resultDf}['{oldColumn}']).dt.days
"""
                else:
                    lines += f"""
# Calculate the difference from the current date
current_date = pd.Timestamp(datetime.now())
{resultDf}['{newColumn}'] = (current_date - {resultDf}['{oldColumn}']).dt.days
"""

                # Uncomment the following block to calculate difference from a specific date:
                # wanted_date = pd.Timestamp(datetime(2024, 2, 29))
                # df[newColumn] = (wanted_date - df[oldColumn]).dt.days

                # Assign the result to the resultDf variable
                return lines, True
            else:
                return "Key Value pair error in dateDiff Function", False
        except Exception as e:
            return str(e), False
        
    def exactVLookup(**kwargs):
        try:
            requiredKeys = ['resultDf', 'leftDf', 'rightDf', 'column', 'columnsFromRight']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                resultDf, leftDf, rightDf, column, columnsFromRight = [kwargs.get(key) for key in requiredKeys]

                column_list = column.split(',')
                for i in range(len(column_list)):
                    column_list[i] = column_list[i].strip()

                columnsFromRightDf_list = columnsFromRight.split(',')
                for i in range(len(columnsFromRightDf_list)):
                    columnsFromRightDf_list[i] = columnsFromRightDf_list[i].strip()
                
                merged_column_list = column_list + columnsFromRightDf_list
                lines = f"""
selected_df = {rightDf}[{merged_column_list}]
{resultDf} = pd.merge({leftDf}, selected_df, on={column_list}, how='left')
"""
                return lines, True
            else:
                return "Key Value pair errors in exact Vlookup", False
        except Exception as e:
            return str(e), False

    def writeExcelPreview(**kwargs):
        try:
            requiredKeys = ['fileName', 'dfSheetNameString', 'temp_dir', 'client', 'recon_id', 'requester_id']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                fileName, dfSheetNameString, temp_dir, client, recon_id, requester_id = [kwargs.get(key) for key in requiredKeys]
                df_list = []
                sheet_name_list = []
                dfAndSheetName = dfSheetNameString.split(', ')
                for pair in dfAndSheetName:
                    df, sheet = pair.split("|")
                    df_list.append(df)
                    sheet_name_list.append(sheet)

                for i in range(len(df_list)):
                    df_list[i] += f"_{client}_{recon_id}_{requester_id}"
                logger.info(df_list)
                df_list = '[' + ', '.join(df_list) + ']'
                fileName_joint = f'{temp_dir}/{fileName}'

                lines = f'''
excel_operation_df = save_dataframes_to_excel({df_list}, {sheet_name_list}, "{fileName_joint}")
'''
                return lines, fileName, True
            else:
                return "Error in key Value Pairs in writeExcel function", False, False
        except Exception as e:
            return str(e), False
        
    
    def if_then_else(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'newColumn', 'otherwiseValue', 'conditions', 'isNumberOtherwiseValue']
            if all(key in kwargs.keys() for key in requiredKeys):
                oldDf, newDf, newColumn, otherwiseValue, conditions, isNumberOtherwiseValue = [kwargs.get(key) for key in requiredKeys]

                # Create condition list for `numpy.select()`
                condition_list = []
                value_list = []

                for condition in conditions:
                    if 'group' in condition:
                        sub_conditions = []
                        for sub_condition in condition['group']:
                            if not sub_condition['isNumber']:
                                value = f"'{sub_condition['value']}'"
                            else:
                                value = f"{sub_condition['value']}"

                            if sub_condition['logicalOperator'] == 'AND':
                                logical_operator = ' & '
                            elif sub_condition['logicalOperator'] == 'OR':
                                logical_operator = ' | '
                            else:
                                logical_operator = ''
                            if sub_condition['operator'] == 'contains':
                                value = f"'{sub_condition['value']}'"
                                sub_condition_line = f'({oldDf}["{sub_condition["columnName"]}"].str.contains({value}, case=False, na=False)){logical_operator}'
                            else:
                                if sub_condition['columnOperation'] != '':
                                    sub_condition_line = f'({oldDf}["{sub_condition["columnName"]}"].str.len() {sub_condition["operator"]} {value}){logical_operator}'
                                else:
                                    sub_condition_line = f'({oldDf}["{sub_condition["columnName"]}"] {sub_condition["operator"]} {value}){logical_operator}'

                            sub_conditions.append(sub_condition_line)


                        condition_line = ''.join(sub_conditions)
                    else:
                        if not condition['isNumber']:
                            value = f'"{condition["value"]}"'
                        else:
                            value = f'{condition["value"]}'

                        if condition['operator'] == 'contains':
                            value = f'"{condition["value"]}"'
                            condition_line = f'{oldDf}["{condition["columnName"]}"].str.contains({value}, case=False, na=False)'
                        else:
                            if condition['columnOperation'] != '':
                                condition_line = f'{oldDf}["{condition["columnName"]}"].str.len() {condition["operator"]} {value}'
                            else:
                                condition_line = f'{oldDf}["{condition["columnName"]}"] {condition["operator"]} {value}'

                    if condition['isNumberOutput']:
                        value_list.append(condition['newValue'])
                    else:
                        value_list.append(f"{condition['newValue']}")

                    condition_list.append(condition_line)
                
                condition_lines_str = '[' + ', '.join(condition_list) + ']'

                if not isNumberOtherwiseValue:
                    lines = f"""
{newDf}['{newColumn}'] = np.select({condition_lines_str}, {value_list}, default = '{otherwiseValue}')
"""
                else:
                    lines = f"""
{newDf}['{newColumn}'] = np.select({condition_lines_str}, {value_list}, default = {otherwiseValue})
"""
                return lines, True
            else:
                return "Kwargs error in if-then-else", False
        except Exception as e:
            return str(e), False
        
    def groupBy(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'groupByColumnString', 'aggregateFunctions', 'verticalSummation', 'horizontalSummation']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newDf, groupByColumnString, aggregateFunctions, verticalSummation, horizontalSummation = [kwargs.get(key) for key in requiredKeys]
                agg_dict = {}
                for function in aggregateFunctions:
                    if function['key'] in agg_dict.keys():
                        agg_list = agg_dict[f"{function['key']}"]
                    else:
                        agg_list = []
                    
                    if function['aggregate'] == 'avg':
                        aggregate_function = ('mean')
                    else:
                        aggregate_function = (function['aggregate'])
                                       
                    if function['value'] == '':
                        agg_list.append(aggregate_function)
                    else:
                        agg_list.append((f"{function['value']}", f"{aggregate_function}"))

                    agg_dict.update({f"{function['key']}": agg_list})

                agg_string = f'agg({agg_dict})'

                groupByColumnlist = groupByColumnString.split(',')
                for i in range(len(groupByColumnlist)):
                    groupByColumnlist[i] = groupByColumnlist[i].strip()

                lines = f"""
{newDf} = {oldDf}.groupby({groupByColumnlist}).{agg_string}
{newDf}.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in {newDf}.columns]
{newDf} = {newDf}.reset_index()
"""
                return lines, True
            else:
                return "key value pair error in pivot", False
        except Exception as e:
            return str(e), False
        
    def pivot(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'groupByColumnString', 'pivotColumn', 'aggregateFunctions', 'horizontalSummation', 'verticalSummation']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newDf, groupByColumnString, pivotColumn, aggregateFunctions, horizontalSummation, verticalSummation = [kwargs.get(key) for key in requiredKeys]
                aggregate_columns = []
                aggregate_functions = []
                for entry in aggregateFunctions:
                    aggregate_columns.append(entry['key'])
                    if entry['aggregate'] == 'avg':
                        aggregate_functions.append('mean')
                    else:
                        aggregate_functions.append(entry['aggregate'])
                aggregate_columns = list(set(aggregate_columns))
                aggregate_functions = list(set(aggregate_functions))

                groupByColumnString = groupByColumnString.split(',')
                for i in range(len(groupByColumnString)):
                    groupByColumnString[i] = groupByColumnString[i].strip()

                             
                if (horizontalSummation & verticalSummation):
                    lines = f"""
# Create a pivot table with more parameters
{newDf} = pd.pivot_table({oldDf}, 
                          values={aggregate_columns},         # Column to aggregate
                          index={groupByColumnString},  # Row index (multiple levels: Date and Region)
                          columns='{pivotColumn.strip()}',      # Columns of the pivot table
                          aggfunc={aggregate_functions},  # Aggregate functions (sum and mean)
                          fill_value=0,           # Replace NaN with 0
                          margins=True,           # Add total row/column (margins)
                          margins_name='Total')   # Name of the total row/column

"""
                else:
                    if horizontalSummation:
                        lines = f"""
# Create a pivot table with more parameters
{newDf} = pd.pivot_table({oldDf}, 
                          values={aggregate_columns},         # Column to aggregate
                          index={groupByColumnString},  # Row index (multiple levels: Date and Region)
                          columns='{pivotColumn.strip()}',      # Columns of the pivot table
                          aggfunc={aggregate_functions},  # Aggregate functions (sum and mean)
                          fill_value=0,           # Replace NaN with 0
                          margins=True,           # Add total row/column (margins)
                          margins_name='Total')   # Name of the total row/column
# Drop the horizontal total row (if present)
{newDf} = {newDf}.drop(index='Total')
"""
                    elif verticalSummation:
                        lines = f"""
# Create a pivot table with more parameters
{newDf} = pd.pivot_table({oldDf}, 
                          values={aggregate_columns},         # Column to aggregate
                          index={groupByColumnString},  # Row index (multiple levels: Date and Region)
                          columns='{pivotColumn.strip()}',      # Columns of the pivot table
                          aggfunc={aggregate_functions},  # Aggregate functions (sum and mean)
                          fill_value=0)           # Replace NaN with 0
# Add column totals (vertical summation)
{newDf}.loc['Total'] = {newDf}.sum()
"""
                    else:
                        lines = f"""
# Create a pivot table with more parameters
{newDf} = pd.pivot_table({oldDf}, 
                          values={aggregate_columns},         # Column to aggregate
                          index={groupByColumnString},  # Row index (multiple levels: Date and Region)
                          columns='{pivotColumn.strip()}',      # Columns of the pivot table
                          aggfunc={aggregate_functions},  # Aggregate functions (sum and mean)
                          fill_value=0)           # Replace NaN with 0
"""
                lines += f"""
{newDf} = {newDf}.reset_index()
"""
                return lines, True
            else:
                return "key value pair error in pivot", False
        except Exception as e:
            return str(e), False

    def getSumOfColumn(**kwargs):
        try:
            requiredKeys = ['df', 'columnName', 'resultDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, columnName, resultDf = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
operation_output = {df}['{columnName}'].sum()
{resultDf} = pd.DataFrame({{'sum': [operation_output]}})
"""
                return lines, True
            else:
                return "Key-value pair error in sum of column", False
        except Exception as e:
            return str(e), False
        
    def getAvgOfColumn(**kwargs):
        try:
            requiredKeys = ['df', 'columnName', 'resultDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, columnName, resultDf = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
operation_output = {df}['{columnName}'].mean()
{resultDf} = pd.DataFrame({{'average': [operation_output]}})
"""
                return lines, True
            else:
                return "Key-value pair error in average of column", False
        except Exception as e:
            return str(e), False
        
    def getMinOfColumn(**kwargs):
        try:
            requiredKeys = ['df', 'columnName', 'resultDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, columnName, resultDf = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
operation_output = {df}['{columnName}'].min()
{resultDf} = pd.DataFrame({{'min': [operation_output]}})
"""
                return lines, True
            else:
                return "Key-value pair error in minimum of column", False
        except Exception as e:
            return str(e), False
        
    def getMaxOfColumn(**kwargs):
        try:
            requiredKeys = ['df', 'columnName', 'resultDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, columnName, resultDf = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
operation_output = {df}['{columnName}'].max()
{resultDf} = pd.DataFrame({{'max': [operation_output]}})
"""
                return lines, True
            else:
                return "Key-value pair error in maximum of column", False
        except Exception as e:
            return str(e), False
        
    def getCountOfColumn(**kwargs):
        try:
            requiredKeys = ['df', 'columnName', 'resultDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, columnName, resultDf = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
operation_output = {df}['{columnName}'].count()
{resultDf} = pd.DataFrame({{'count': [operation_output]}})
"""
                return lines, True
            else:
                return "Key-value pair error in count of column", False
        except Exception as e:
            return str(e), False
        
    def addColumns(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newColumn', 'columnString']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newColumn, columnString = [kwargs.get(key) for key in requiredKeys]
                columnsToAdd = columnString.split(',')
                for i in range(len(columnsToAdd)):
                    columnsToAdd[i] = columnsToAdd[i].strip()
                columnstr = ' + '.join([f"{oldDf}['{column}']" for column in columnsToAdd])

                lines = f"""
{oldDf}['{newColumn}'] = {columnstr}
"""
                return lines, True
            else:
                return "Key-value pair error in addColumns function", False
        except Exception as e:
            return str(e), False
        
    def subtractColumns(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newColumn', 'columnString']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newColumn, columnString = [kwargs.get(key) for key in requiredKeys]
                columnsToSubtract = columnString.split(',')
                for i in range(len(columnsToSubtract)):
                    columnsToSubtract[i] = columnsToSubtract[i].strip()
                columnstr = ' - '.join([f"{oldDf}['{column}']" for column in columnsToSubtract])

                lines = f"""
{oldDf}['{newColumn}'] = {columnstr}
"""
                return lines, True
            else:
                return "Key-value pair error in subtractColumns function", False
        except Exception as e:
            return str(e), False
        
    def multiplyColumns(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newColumn', 'columnString']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newColumn, columnString = [kwargs.get(key) for key in requiredKeys]
                columnsToMultiply = columnString.split(',')
                for i in range(len(columnsToMultiply)):
                    columnsToMultiply[i] = columnsToMultiply[i].strip()
                columnstr = ' * '.join([f"{oldDf}['{column}']" for column in columnsToMultiply])

                lines = f"""
{oldDf}['{newColumn}'] = {columnstr}
"""
                return lines, True
            else:
                return "Key-value pair error in multiplyColumns function", False
        except Exception as e:
            return str(e), False
        
    def divideColumns(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newColumn', 'columnString']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newColumn, columnString = [kwargs.get(key) for key in requiredKeys]
                columnsToDivide = columnString.split(',')
                for i in range(len(columnsToDivide)):
                    columnsToDivide[i] = columnsToDivide[i].strip()
                columnstr = ' / '.join([f"{oldDf}['{column}']" for column in columnsToDivide])
                

                lines = f"""
{oldDf}['{newColumn}'] = {columnstr}
"""
                return lines, True
            else:
                return "Key-value pair error in divideColumns function", False
        except Exception as e:
            return str(e), False
        
    def findAndReplaceInColumn(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'configList']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newDf, findandreplacelist = [kwargs.get(key) for key in requiredKeys]
                
                lines = f"""{newDf} = {oldDf}.copy()\n"""
                for entry in findandreplacelist:
                    lines += f"""{newDf}["{entry['newColumn']}"] = {newDf}["{entry['oldColumn']}"].replace("{entry['oldValue']}", "{entry['newValue']}")\n"""
                return lines, True
            else:
                return "Key-value pair error in findAndReplaceInColumn function", False
        except Exception as e:
            return str(e), False
        
    def sortDf(**kwargs):
        '''
        Takes 3 arguments:
        1. old Table Name (DataFrame).
        2. new Table name (DataFrame).
        3. column dict list (list of dictionaries):
            [
                {'columnName': <columnName>,
                'sortType': <'desc'/'asc'>},
                {'columnName': <columnName>,
                'sortType': <'desc'/'asc'>}
            ]
        '''
        try:
            requiredKeys = ['oldDf', 'newDf', 'columnDictList']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newDf, columnDictList = [kwargs.get(key) for key in requiredKeys]
                by_list = []
                asc_list = []
                for entry in columnDictList:
                    by_list.append(entry['columnName'])
                    if entry['sortType'] == 'asc':
                        asc_list.append(True)
                    else:
                        asc_list.append(False)
                lines = f"""
{newDf} = {oldDf}.sort_values(by = {by_list}, ascending = {asc_list})
"""
                return lines, True
            else:
                return "Key-value pair error in sortDf function", False
        except Exception as e:
            return str(e), False
        
    def combineTables(**kwargs):
        try:
            requiredKeys = ['resultDf', 'inputDataframeList', 'client', 'recon_id', 'requester_id']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                resultDf, inputDataframeList, client, recon_id, requester_id= [kwargs.get(key) for key in requiredKeys]
                
                """
                Merges multiple DataFrames one below the other in pandas.
                
                Parameters:
                inputDataframeList (List[pd.DataFrame]): List of pandas DataFrames to be merged.

                Returns:
                DataFrame: A single DataFrame that is the result of merging all input DataFrames.
                """
                for i in range(len(inputDataframeList)):
                    inputDataframeList[i] += f"_{client}_{recon_id}_{requester_id}"
                input_list = '[' + ', '.join(inputDataframeList) + ']'
                lines = f"""
# Merge all DataFrames in the list using pd.concat
{resultDf} = pd.concat({input_list}, ignore_index=True)
"""
                return lines, True
            else:
                return "Key-value pair error in combineTables function", False
        except Exception as e:
            return str(e), False
        
    def mid(**kwargs):
        try:
            requiredKeys = ['df', 'oldColumn', 'newColumn', 'startPosition', 'buffer']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, oldColumn, newColumn,startPosition, buffer = [kwargs.get(key) for key in requiredKeys]
                lines = f"{df}['{newColumn}'] = {df}['{oldColumn}'].str[{int(startPosition) - 1}:{int(startPosition) + int(buffer) - 1}]"
                return lines, True
            else:
                return "Key-value pair error in mid function", False
        except Exception as e:
            return f"Error in PreviewRepoService mid() {str(e)}", False