# from app.services.recon.recon_mongo import DocumentService
# from pyspark.sql import SparkSession
from app.helpers.common import sync_dir_structure
from app.settings import *
from app.constants.ADLS import *
import os
from app.helpers.common import *

# sync_dir_structure(JAR_DIR)
jar_dir = JAR_DIR
# sync_dir_structure(TRANSFORMATION_OUTPUT_DIR)
output_dir = TRANSFORMATION_OUTPUT_DIR
template_dir = TRANSFORMATION_TEMPLATEFILES_DIR

class RepoService:
    def imports():
        '''
        This function returns the import structure for every python file to be created.

        '''
        lines = f"""
from pyspark.sql import SparkSession
from pyspark.sql.functions import *
from openpyxl import load_workbook
from pyspark.sql.types import *
import functools
import builtins
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pyspark.sql.window import *
import numpy as np
import logging
from functools import reduce
from operator import add
from datetime import datetime
import builtins
import paramiko
from io import StringIO
import tempfile

def getSftpCon(hostname, port, username,password = None, pkey = None):
    transport = paramiko.Transport(hostname, int(port))

    if pkey:
        key = paramiko.RSAKey(file_obj=StringIO(pkey.strip()))
        transport.connect(username = username, pkey=key)
    else:
        transport.connect(username=username, password=password)
    sftp = paramiko.SFTPClient.from_transport(transport)
    return sftp

def save_dataframes_to_excel(dataframes, sheet_names, excel_filename):
    '''
    Save multiple PySpark DataFrames into a single Excel file with multiple sheets.

    :param dataframes: List of PySpark DataFrames
    :param sheet_names: List of sheet names corresponding to each DataFrame
    :param excel_filename: Name of the Excel file to create
    '''
    try:
        # Load the existing workbook if it exists
        workbook = load_workbook(filename=excel_filename)
    except FileNotFoundError:
        # If the workbook doesn't exist, create a new one
        workbook = openpyxl.Workbook()

    # Add each DataFrame to the Excel workbook as a separate sheet
    for df, sheet_name in zip(dataframes, sheet_names):
        # Check if the sheet already exists
        # Create a new sheet with the given name
        worksheet = workbook.create_sheet(title=sheet_name)
        # Collect the data from the DataFrame
        data = df.collect()  # Collect the data locally
        # Convert the data into rows that can be added to the Excel sheet
        # Get column names
        col_names = df.columns
        # Add column headers to the sheet
        worksheet.append(col_names)
        # Add data rows to the sheet
        for row in data:
            worksheet.append(row)

    # Save the workbook to the specified file
    workbook.save(excel_filename)

################### creating a spark Session spark session ###########################
spark = SparkSession.builder \\
                    .appName("spark session") \\
                    .config("spark.jars", "{jar_dir.strip()}/spark-excel_2.13-3.3.1_0.18.5.jar") \\
                    .config("spark.driver.extraJavaOptions", "-Dlog4j.configuration=log4j.properties") \\
                    .getOrCreate()
spark.conf.set("spark.sql.decimalOperations.allowPrecisionLoss", "false")
sc = spark.sparkContext
sc.setLogLevel("ERROR")
#spark.conf.set("spark.sql.repl.eagerEval.enabled", False)
"""
        return lines
    
    def imports_transformation():
        '''
        This function returns the import structure for every python file to be created.

        '''
        # .set("spark.jars", jars)
        # .set("spark.jars.packages", "org.apache.iceberg:iceberg-spark-runtime-3.3_2.13:1.5.2,org.apache.hadoop:hadoop-azure:3.3.6,org.apache.hadoop:hadoop-common:3.3.6,org.apache.iceberg:iceberg-hive-runtime:1.5.2,org.wildfly.openssl:wildfly-openssl:2.2.5.Final,com.crealytics:spark-excel_2.12:3.5.1_0.20.4")
        # .set('spark.sql.extensions', "org.apache.iceberg.spark.extensions.IcebergSparkSessionExtensions")
        # .set("spark.sql.catalog.r2rdevpoccatalog", "org.apache.iceberg.spark.SparkCatalog")
        # .set("spark.sql.catalog.r2rdevpoccatalog.type","hive")
        # .set("spark.sql.catalog.r2rdevpoccatalog.uri","thrift://localhost:9083")
        # .set("hive.metastore.uris", "thrift://localhost:9083")

        lines = f"""from pyspark.sql import SparkSession
from pyspark.sql.functions import *
from openpyxl import load_workbook
from pyspark.sql.types import *
import functools
import builtins
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pyspark.sql.window import *
import numpy as np
import logging
from functools import reduce
from operator import add
from datetime import datetime
import builtins
import pyspark
import paramiko
from fastapi.responses import JSONResponse
from io import StringIO
import tempfile
import sys
import os
from typing import Optional
import zipfile
import shutil
from io import BytesIO
from openpyxl import Workbook
from azure.storage.filedatalake import DataLakeServiceClient
from azure.storage.blob import BlobServiceClient
import requests
import json



# sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# from app.services.recon.recon_mongo import DocumentService
# from app.models.recon.document_model_mongo import DocumentModel
# from app.settings import *
# from app.helpers.recon import get_sftp
# from app.models.recon import document_model_mongo
# from app.helpers.transformation import *
# from app.helpers.common import *
# from app.helpers.preview import *
"""
        return lines
    def getFilePathOnADLS():
        lines = f"""
def getFilePathOnADLS(directory,
                      contains=None,
                      startsWith=None,
                      endsWith=None,
                      equals=None):
    try:
        matching_files = []
        account_url = f"https://{{os.getenv('AZURE_STORAGE_ACCOUNT_NAME')}}.dfs.core.windows.net"
        credential = os.getenv("AZURE_STORAGE_ACCOUNT_KEY")
        container_name = os.getenv("AZURE_BLOB_CONTAINER_NAME")

        service_client = DataLakeServiceClient(account_url=account_url, credential=credential)

        # Get the file system client (which is equivalent to a container in ADLS Gen2)
        file_system_client = service_client.get_file_system_client(file_system=container_name)

        # Walk through the directory and its subdirectories
        paths = file_system_client.get_paths(path=directory, recursive=True)

        for path in paths:
            filename = path.name.split('/')[-1]  # Extract the filename
            if filename.startswith("._") or filename.startswith("_") or filename.startswith("_MACOSX"):
                continue
            else:
                # Ignore snapshot files, files without an extension, and .zip files
                if '.' in filename[1:]:
                    if filename.endswith(".zip"):
                        pass
                    elif contains and contains in filename:
                        matching_files.append(path.name)
                    elif startsWith and filename.startswith(startsWith):
                        matching_files.append(path.name)
                    elif endsWith and filename.endswith(endsWith):
                        matching_files.append(path.name)
                    elif equals and filename == equals:
                        matching_files.append(path.name)

        if len(matching_files) > 0:
            return matching_files[0]
        else:
            return None

    except Exception as e:
        pass
"""
        return lines
    
    def zipFilesOnADLS():
        lines = f"""
def zipFilesOnADLS(file_paths, zip_file_path):
    # Initialize the connection to ADLS Gen2
    account_url = f"https://{{os.getenv('AZURE_STORAGE_ACCOUNT_NAME')}}.dfs.core.windows.net"
    credential = os.getenv("AZURE_STORAGE_ACCOUNT_KEY")
    container_name = os.getenv("AZURE_BLOB_CONTAINER_NAME")

    service_client = DataLakeServiceClient(account_url=account_url, credential=credential)

    filesystem_client = service_client.get_file_system_client(container_name)

    # Create an in-memory BytesIO stream for the zip file
    zip_stream = BytesIO()

    # Create a ZipFile object and add files
    with zipfile.ZipFile(zip_stream, 'w') as zip_file:
        for file_path in file_paths:
            print(file_path)
            # Download file content
            file_client = filesystem_client.get_file_client(file_path)
            download = file_client.download_file()
            file_content = download.readall()

            # Add the file to the zip (write the file with the same name as in ADLS)
            zip_file.writestr(file_path.split("/")[-1], file_content)

    # After writing to the zip file, seek back to the start of the stream
    zip_stream.seek(0)

    # Upload the zip file back to ADLS Gen2
    zip_file_client = filesystem_client.get_file_client(zip_file_path)
    zip_file_client.upload_data(zip_stream, overwrite=True)

    print("Zip file created and uploaded successfully.")
"""
        return lines
    
    def sftpConn():
        lines = """
def sftp_mkdirs(sftp, remote_path, mode=511):
    dirs = remote_path.split('/')
    path = ''
    for directory in dirs:
        if directory:  # to handle leading '/'
            path += '/' + directory
            try:
                sftp.mkdir(path, mode=mode)
            except IOError:  # Directory already exists
                pass
                
def get_sftp(hostname: str,
    port: str,
    username: str,
    password: Optional[str] = None,
    key_file: Optional[str] = None):

    try:
        if not password and not key_file:
            raise JSONResponse(status_code=400, detail="Either password or key_file must be provided")
        else:
            transport = paramiko.Transport((hostname, port))
            if key_file is not None:
                private_key_file = StringIO(key_file.strip())
                try:
                    private_key = paramiko.RSAKey.from_private_key(private_key_file)
                except paramiko.ssh_exception.SSHException:
                    private_key_file.seek(0)
                    try:
                        private_key = paramiko.ECDSAKey.from_private_key(private_key_file)
                    except paramiko.ssh_exception.SSHException:
                        private_key_file.seek(0)
                        private_key = paramiko.Ed25519Key.from_private_key(private_key_file)

                    # Connect to the SFTP server
                try:
                    transport.connect(username=username, pkey=private_key)
                except paramiko.SSHException as e:
                    print(f"SSH connection failed: {e}")
            else:
                transport.connect(username=username, password=password)

            sftp = paramiko.SFTPClient.from_transport(transport)
            print("created sftp")
            return sftp
        
    except Exception as e:
        print(f"Connection failed: {str(e)}")
        return JSONResponse(status_code=500,
                            content={"message": f"Connection failed: {str(e)}"})
"""
        return lines
    
    def saveDFToExcel():
        lines = f"""
def save_dataframes_to_excel(dataframes, sheet_names, adls_file_path, file_system_name = os.getenv("AZURE_BLOB_CONTAINER_NAME")):
    '''
    Save multiple PySpark DataFrames into a single Excel file with multiple sheets and upload to ADLS Gen2 from Databricks.

    :param dataframes: List of PySpark DataFrames
    :param sheet_names: List of sheet names corresponding to each DataFrame
    :param adls_account_name: ADLS Gen2 account name
    :param file_system_name: The name of the ADLS Gen2 filesystem
    :param remote_file_path: Path to save the file in ADLS Gen2
    :param client_id: Service principal client ID
    :param tenant_id: Azure tenant ID
    :param client_secret: Service principal client secret
    '''
    
    # Create a BytesIO stream to save the Excel file in memory
    excel_stream = BytesIO()

    # Create a new workbook
    workbook = Workbook()
    
    # Remove the default sheet created in the workbook
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Add each DataFrame to the Excel workbook as a separate sheet
    for df, sheet_name in zip(dataframes, sheet_names):
        # Create a new sheet with the given name
        worksheet = workbook.create_sheet(title=sheet_name)
        # Collect the data from the DataFrame
        data = df.collect()  # Collect the data locally
        # Get column names
        col_names = df.columns
        # Add column headers to the sheet
        worksheet.append(col_names)
        # Add data rows to the sheet
        for row in data:
            worksheet.append(list(row))

    # Save the workbook to the BytesIO stream
    workbook.save(excel_stream)
    excel_stream.seek(0)  # Reset the stream position to the beginning

    # Authenticate with ADLS Gen2 using the service principal
    service_client = DataLakeServiceClient(
        account_url = f"https://{{os.getenv('AZURE_STORAGE_ACCOUNT_NAME')}}.dfs.core.windows.net",
        credential=os.getenv("AZURE_STORAGE_ACCOUNT_KEY")
    )
    
    # Get the file system client
    file_system_client = service_client.get_file_system_client(file_system_name)
    
    # Create or get the directory client where the file will be saved
    directory_client = file_system_client.get_directory_client(adls_file_path.rsplit("/", 1)[0])
    
    # Create or get the file client for the Excel file
    file_client = directory_client.create_file(adls_file_path.rsplit("/", 1)[1])

    # Upload the in-memory Excel file to ADLS Gen2
    file_client.upload_data(excel_stream, overwrite=True)

    print(f"File uploaded to ADLS Gen2: {{adls_file_path}}")
"""
        return lines
    
    def startingBlock():
        lines = f"""
try:
    conf = (
    pyspark.SparkConf().setMaster("local")
        .setAppName('app_name')
        .set(f"fs.azure.account.key.{{os.getenv('AZURE_STORAGE_ACCOUNT_NAME')}}.dfs.core.windows.net", os.getenv('AZURE_STORAGE_ACCOUNT_KEY'))
        )

    spark = SparkSession.builder.config(conf=conf).getOrCreate()
"""
        return lines

    def readExcel(**kwargs):
        try:
            requiredKeys = ['sheetName', 'df', 'localFilePath', 'fileName']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for key, value in zip(requiredKeys, values) if key != 'fileName'):
                    sheetName, df, localFilePath, fileName = values
                    if fileName == '':
                        fileName = ''
                    else:
                        fileName = fileName.split('/')[-1]
                        
                    lines = f"""
####### read excel file into pyspark ##########
{df} = spark.read \\
            .format("com.crealytics.spark.excel") \\
            .option("header", "true") \\
            .option("inferSchema", "true") \\
            .option("sheetName", "{sheetName}") \\
            .load(f"abfss://{{os.getenv('AZURE_BLOB_CONTAINER_NAME')}}@{{os.getenv('AZURE_STORAGE_ACCOUNT_NAME')}}.dfs.core.windows.net/{localFilePath.strip()}/{fileName}")

for cols in {df}.columns:
    new_col = cols.replace(".", "")  # Remove all periods from the column name
    {df} = {df}.withColumnRenamed(cols, new_col)

columns_to_keep = [col for col in {df}.columns if not col.startswith("_c")]
# Create a new DataFrame with only the columns to retain
{df} = {df}.select(columns_to_keep)
{df}.show()
"""
                    return lines, True
                else:
                    return "filePath, fileName or sheetName not provided", False
                
            else:
                return "filePath or sheetName not provided", False
        
        except Exception as e:
            return str(e), False

    def readCsv(**kwargs):
        try:
            requiredKeys = ['fileName', 'delimiter', 'df', 'localFilePath']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for key, value in zip(requiredKeys, values) if key != 'fileName'):
                    fileName, delimiter, df, localFilePath = values
                    if fileName == '':
                        fileName = ''
                    else:
                        fileName = fileName.split('/')[-1]
                    lines = f"""   
############ read CSV files ############
{df} = spark.read.options(delimiter = '{delimiter}',header = True, inferSchema = True).csv(f"abfss://{{os.getenv('AZURE_BLOB_CONTAINER_NAME')}}@{{os.getenv('AZURE_STORAGE_ACCOUNT_NAME')}}.dfs.core.windows.net/{localFilePath.strip()}/{fileName}")
for cols in {df}.columns:
    new_col = cols.replace(".", "")  # Remove all periods from the column name
    {df} = {df}.withColumnRenamed(cols, new_col)

{df}.show()
"""
                    return lines, True
                else:
                    return "filePath not provided", False
            else:
                return "filePath or fileName not provided", False
        
        except Exception as e:
            return str(e), False

    def selectCols(dropna, **kwargs):
        try:
            newDf_found = False
            oldDf_found = False
            columnString_found = False

            for key, value in kwargs.items():
                if key == 'newDf':
                    newDf_found = True
                    newDf = value
                if key == 'oldDf':
                    oldDf_found = True
                    oldDf = value
                if key == 'columnString':
                    columnString_found = True
                    columnString = value

            if columnString_found:
                col_list = columnString.split(',')
                col_list = list(map(lambda x: x.strip(), col_list))

                columnString = ''
                for col in col_list:
                    if ' as ' in col:
                        col = col.split(' as ')
                        columnString += f"col('{col[0]}').alias('{col[1]}'), "
                    else:
                        columnString += f"col('{col.strip()}'), "
                
                columnString = columnString[:-2]
                
                if (oldDf_found & newDf_found):
                    if dropna:
                        pyspark_command = f"""   
{newDf} = {oldDf}.select({columnString}).dropna()
{newDf}.show()
"""
                    else:
                        pyspark_command = f"""  
{newDf} = {oldDf}.select({columnString})
{newDf}.show()
"""
                    return pyspark_command, True
                else:
                    return "Key Value pair errors", False
            else:
                return "Key Value pair errors selectCols function", False
        
        except Exception as e:
            return str(e), False

    def concatWs(**kwargs):
        try:
            requiredKeys = ['delimiter', 'columnString','df','newCol']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                delimiter, columnString, df, newCol = values
                columnString = columnString.split(',')
                column_list = list(map(lambda x: x.strip(), columnString))

                prep_cols = ""
                for col in column_list:
                    prep_cols += f"col('{col}').cast('string'), "
                prep_cols = prep_cols[:-2]
                pyspark_command = f"""
################ command for concat_ws ###################
{df} = {df}.withColumn('{newCol}', concat_ws('{delimiter}', {prep_cols}))
{df}.show()
"""
                return pyspark_command, True
            else:
                return "Key Value pair errors in concatWs function", False
            
        except Exception as e:
            return str(e), False

    def filterDf(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'filterConditions']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    oldDf, newDf, filterConditions = values
                    operationLines = f"{newDf} = {oldDf}"
                    for condition in filterConditions:
                        if 'group' not in condition:
                            columnOperation = condition.get('columnOperation')
                            column = condition.get('columnName')
                            operator = condition.get('operator')
                            value = condition.get('value')
                            isValueNumeric = condition.get('isNumber')
                            if columnOperation == '':
                                if operator == '>':
                                    if isValueNumeric:
                                        operationLines += f".filter(col('{column}') > {value})"
                                    else:
                                        operationLines += f".filter(col('{column}') > '{value}')"
                                elif operator == '<':
                                    if isValueNumeric:
                                        operationLines += f".filter(col('{column}') < {value})"
                                    else:
                                        operationLines += f".filter(col('{column}') < '{value}')"
                                elif operator == '==':
                                    if isValueNumeric:
                                        operationLines += f".filter(col('{column}') == {value})"
                                    else:
                                        operationLines += f".filter(col('{column}') == '{value}')"
                                elif operator == '!=':
                                    if isValueNumeric:
                                        operationLines += f".filter(col('{column}') != {value})"
                                    else:
                                        operationLines += f".filter(col('{column}') != '{value}')"
                                elif operator == 'like':
                                    operationLines += f".filter(col('{column}').like('{value}'))"
                                elif operator == 'rlike':
                                    operationLines += f".filter(col('{column}').rlike('{value}'))"
                                elif operator == 'isNotNull':
                                    operationLines += f".filter(col('{column}').isNotNull())"
                                elif operator == 'isNull':
                                    operationLines += f".filter(col('{column}').isNull())"
                                elif operator == 'contains':
                                    operationLines += f".filter(col('{column}').contains('{value}'))"
                                else:
                                    operationLines = operationLines
                            else:
                                if operator == '>':
                                    if isValueNumeric:
                                        operationLines += f".filter({columnOperation}(col('{column}')) > {value})"
                                    else:
                                        operationLines += f".filter({columnOperation}(col('{column}')) > '{value}')"
                                elif operator == '<':
                                    if isValueNumeric:
                                        operationLines += f".filter({columnOperation}(col('{column}')) < {value})"
                                    else:
                                        operationLines += f".filter({columnOperation}(col('{column}')) < '{value}')"
                                elif operator == '==':
                                    if isValueNumeric:
                                        operationLines += f".filter({columnOperation}(col('{column}')) == {value})"
                                    else:
                                        operationLines += f".filter({columnOperation}(col('{column}')) == '{value}')"
                                elif operator == '!=':
                                    if isValueNumeric:
                                        operationLines += f".filter({columnOperation}(col('{column}')) != {value})"
                                    else:
                                        operationLines += f".filter({columnOperation}(col('{column}')) != '{value}')"
                                elif operator == 'like':
                                    operationLines += f".filter(col('{column}').like('{value}'))"
                                elif operator == 'rlike':
                                    operationLines += f".filter(col('{column}').rlike('{value}'))"
                                elif operator == 'isNotNull':
                                    operationLines += f".filter(col('{column}').isNotNull())"
                                elif operator == 'isNull':
                                    operationLines += f".filter(col('{column}').isNull())"
                                elif operator == 'contains':
                                    operationLines += f".filter(col('{column}').contains('{value}'))"
                                else:
                                    operationLines = operationLines
                        else:
                            sub_condition_line = ''
                            for sub_condition in condition['group']:
                                columnOperation = sub_condition.get('columnOperation')
                                column = sub_condition.get('columnName')
                                operator = sub_condition.get('operator')
                                value = sub_condition.get('value')
                                isValueNumeric = sub_condition.get('isNumber')
                                logicalOperator = sub_condition.get('logicalOperator')

                                if logicalOperator == 'AND':
                                    logical_operator = ' & '
                                elif logicalOperator == 'OR':
                                    logical_operator = ' | '
                                else:
                                    logical_operator = ''
                                if columnOperation == '':     
                                    if operator == '>':
                                        if isValueNumeric:
                                            sub_condition_line += f"(col('{column}') > {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"(col('{column}') > '{value}') {logical_operator}"
                                    elif operator == '<':
                                        if isValueNumeric:
                                            sub_condition_line += f"(col('{column}') < {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"(col('{column}') < '{value}') {logical_operator}"
                                    elif operator == '==':
                                        if isValueNumeric:
                                            sub_condition_line += f"(col('{column}') == {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"(col('{column}') == '{value}') {logical_operator}"
                                    elif operator == '!=':
                                        if isValueNumeric:
                                            sub_condition_line += f"(col('{column}') != {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"(col('{column}') != '{value}') {logical_operator}"
                                    elif operator == 'like':
                                        sub_condition_line += f"(col('{column}').like('{value}')) {logical_operator}"
                                    elif operator == 'rlike':
                                        sub_condition_line += f"(col('{column}').rlike('{value}')) {logical_operator}"
                                    elif operator == 'isNotNull':
                                        sub_condition_line += f"(col('{column}').isNotNull()) {logical_operator}"
                                    elif operator == 'isNull':
                                        sub_condition_line += f"(col('{column}').isNull()) {logical_operator}"
                                    elif operator == 'contains':
                                        sub_condition_line += f"(col('{column}').contains('{value}')) {logical_operator}"
                                    else:
                                        sub_condition_line = sub_condition_line
                                else:
                                    if operator == '>':
                                        if isValueNumeric:
                                            sub_condition_line += f"({columnOperation}(col('{column}')) > {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"({columnOperation}(col('{column}')) > '{value}') {logical_operator}"
                                    elif operator == '<':
                                        if isValueNumeric:
                                            sub_condition_line += f"({columnOperation}(col('{column}')) < {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"({columnOperation}(col('{column}')) < '{value}') {logical_operator}"
                                    elif operator == '==':
                                        if isValueNumeric:
                                            sub_condition_line += f"({columnOperation}(col('{column}')) == {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"({columnOperation}(col('{column}')) == '{value}') {logical_operator}"
                                    elif operator == '!=':
                                        if isValueNumeric:
                                            sub_condition_line += f"({columnOperation}(col('{column}')) != {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"({columnOperation}(col('{column}')) != '{value}') {logical_operator}"
                                    elif operator == 'like':
                                        sub_condition_line += f"(col('{column}').like('{value}')) {logical_operator}"
                                    elif operator == 'rlike':
                                        sub_condition_line += f"(col('{column}').rlike('{value}')) {logical_operator}"
                                    elif operator == 'isNotNull':
                                        sub_condition_line += f"(col('{column}').isNotNull()) {logical_operator}"
                                    elif operator == 'isNull':
                                        sub_condition_line += f"(col('{column}').isNull()) {logical_operator}"
                                    elif operator == 'contains':
                                        sub_condition_line += f"(col('{column}').contains('{value}')) {logical_operator}"
                                    else:
                                        sub_condition_line = sub_condition_line

                            sub_condition_line = f".filter({sub_condition_line.strip()})"
                            operationLines += sub_condition_line

                    lines = f"""
{operationLines}
{newDf}.show()
"""
                    return lines, True
                else:
                    return "Key Value pair errors in filterDf errors", False
            else:
                return "Key Value pair errors in filterDf errors", False
        except Exception as e:
            return str(e), False
        
    def filterDelete(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'filterConditions']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    oldDf, newDf, filterConditions = values
                    operationLines = f"{newDf} = {oldDf}"
                    for condition in filterConditions:
                        if 'group' not in condition:
                            columnOperation = condition.get('columnOperation')
                            column = condition.get('columnName')
                            operator = condition.get('operator')
                            value = condition.get('value')
                            isValueNumeric = condition.get('isNumber')
                            if columnOperation == '':
                                if operator == '>':
                                    if isValueNumeric:
                                        operationLines += f".filter(~(col('{column}') > {value}))"
                                    else:
                                        operationLines += f".filter(~(col('{column}') > '{value}'))"
                                elif operator == '<':
                                    if isValueNumeric:
                                        operationLines += f".filter(~(col('{column}') < {value}))"
                                    else:
                                        operationLines += f".filter(~(col('{column}') < '{value}'))"
                                elif operator == '==':
                                    if isValueNumeric:
                                        operationLines += f".filter(~(col('{column}') == {value}))"
                                    else:
                                        operationLines += f".filter(~(col('{column}') == '{value}'))"
                                elif operator == '!=':
                                    if isValueNumeric:
                                        operationLines += f".filter(~(col('{column}') != {value}))"
                                    else:
                                        operationLines += f".filter(~(col('{column}') != '{value}'))"
                                elif operator == 'like':
                                    operationLines += f".filter(~(col('{column}').like('{value}')))"
                                elif operator == 'rlike':
                                    operationLines += f".filter(~(col('{column}').rlike('{value}')))"
                                elif operator == 'isNotNull':
                                    operationLines += f".filter(~(col('{column}').isNotNull()))"
                                elif operator == 'isNull':
                                    operationLines += f".filter(~(col('{column}').isNull()))"
                                elif operator == 'contains':
                                    operationLines += f".filter(~(col('{column}').contains('{value}')))"
                                else:
                                    operationLines = operationLines
                            else:
                                if operator == '>':
                                    if isValueNumeric:
                                        operationLines += f".filter(~({columnOperation}(col('{column}')) > {value}))"
                                    else:
                                        operationLines += f".filter(~({columnOperation}(col('{column}')) > '{value}'))"
                                elif operator == '<':
                                    if isValueNumeric:
                                        operationLines += f".filter(~({columnOperation}(col('{column}')) < {value}))"
                                    else:
                                        operationLines += f".filter(~({columnOperation}(col('{column}')) < '{value}'))"
                                elif operator == '==':
                                    if isValueNumeric:
                                        operationLines += f".filter(~({columnOperation}(col('{column}')) == {value}))"
                                    else:
                                        operationLines += f".filter(~({columnOperation}(col('{column}')) == '{value}'))"
                                elif operator == '!=':
                                    if isValueNumeric:
                                        operationLines += f".filter(~({columnOperation}(col('{column}')) != {value}))"
                                    else:
                                        operationLines += f".filter(~({columnOperation}(col('{column}')) != '{value}'))"
                                elif operator == 'like':
                                    operationLines += f".filter(~(col('{column}').like('{value}')))"
                                elif operator == 'rlike':
                                    operationLines += f".filter(~(col('{column}').rlike('{value}')))"
                                elif operator == 'isNotNull':
                                    operationLines += f".filter(~(col('{column}').isNotNull()))"
                                elif operator == 'isNull':
                                    operationLines += f".filter(~(col('{column}').isNull()))"
                                elif operator == 'contains':
                                    operationLines += f".filter(~(col('{column}').contains('{value}')))"
                                else:
                                    operationLines = operationLines
                        else:
                            sub_condition_line = ''
                            for sub_condition in condition['group']:
                                columnOperation = sub_condition.get('columnOperation')
                                column = sub_condition.get('columnName')
                                operator = sub_condition.get('operator')
                                value = sub_condition.get('value')
                                isValueNumeric = sub_condition.get('isNumber')
                                logicalOperator = sub_condition.get('logicalOperator')

                                if logicalOperator == 'AND':
                                    logical_operator = ' & '
                                elif logicalOperator == 'OR':
                                    logical_operator = ' | '
                                else:
                                    logical_operator = ''
                                if columnOperation == '':     
                                    if operator == '>':
                                        if isValueNumeric:
                                            sub_condition_line += f"~(col('{column}') > {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"~(col('{column}') > '{value}') {logical_operator}"
                                    elif operator == '<':
                                        if isValueNumeric:
                                            sub_condition_line += f"~(col('{column}') < {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"~(col('{column}') < '{value}') {logical_operator}"
                                    elif operator == '==':
                                        if isValueNumeric:
                                            sub_condition_line += f"~(col('{column}') == {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"~(col('{column}') == '{value}') {logical_operator}"
                                    elif operator == '!=':
                                        if isValueNumeric:
                                            sub_condition_line += f"~(col('{column}') != {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"~(col('{column}') != '{value}') {logical_operator}"
                                    elif operator == 'like':
                                        sub_condition_line += f"~(col('{column}').like('{value}')) {logical_operator}"
                                    elif operator == 'rlike':
                                        sub_condition_line += f"~(col('{column}').rlike('{value}')) {logical_operator}"
                                    elif operator == 'isNotNull':
                                        sub_condition_line += f"~(col('{column}').isNotNull()) {logical_operator}"
                                    elif operator == 'isNull':
                                        sub_condition_line += f"~(col('{column}').isNull()) {logical_operator}"
                                    elif operator == 'contains':
                                        sub_condition_line += f"~(col('{column}').contains('{value}')) {logical_operator}"
                                    else:
                                        sub_condition_line = sub_condition_line
                                else:
                                    if operator == '>':
                                        if isValueNumeric:
                                            sub_condition_line += f"~({columnOperation}(col('{column}')) > {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"~({columnOperation}(col('{column}')) > '{value}') {logical_operator}"
                                    elif operator == '<':
                                        if isValueNumeric:
                                            sub_condition_line += f"~({columnOperation}(col('{column}')) < {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"~({columnOperation}(col('{column}')) < '{value}') {logical_operator}"
                                    elif operator == '==':
                                        if isValueNumeric:
                                            sub_condition_line += f"~({columnOperation}(col('{column}')) == {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"~({columnOperation}(col('{column}')) == '{value}') {logical_operator}"
                                    elif operator == '!=':
                                        if isValueNumeric:
                                            sub_condition_line += f"~({columnOperation}(col('{column}')) != {value}) {logical_operator}"
                                        else:
                                            sub_condition_line += f"~({columnOperation}(col('{column}')) != '{value}') {logical_operator}"
                                    elif operator == 'like':
                                        sub_condition_line += f"~(col('{column}').like('{value}')) {logical_operator}"
                                    elif operator == 'rlike':
                                        sub_condition_line += f"~(col('{column}').rlike('{value}')) {logical_operator}"
                                    elif operator == 'isNotNull':
                                        sub_condition_line += f"~(col('{column}').isNotNull()) {logical_operator}"
                                    elif operator == 'isNull':
                                        sub_condition_line += f"~(col('{column}').isNull()) {logical_operator}"
                                    elif operator == 'contains':
                                        sub_condition_line += f"~(col('{column}').contains('{value}')) {logical_operator}"
                                    else:
                                        sub_condition_line = sub_condition_line

                            sub_condition_line = f".filter({sub_condition_line.strip()})"
                            operationLines += sub_condition_line

                    lines = f"""
{operationLines}
{newDf}.show()
"""
                    return lines, True
                else:
                    return "Key Value pair errors in filterDelete() errors", False
            else:
                return "Key Value pair errors in filterDelete() errors", False
        except Exception as e:
            return str(e), False

    def renameColumns(**kwargs):
        try:
            requiredKeys = ['df', 'renameColumnString']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    df, renameColumnString = values
                    renameColumnString = renameColumnString.split(',')
                    renameColumnString = list(map(lambda x: x.strip(), renameColumnString))
                    renamed_cols = {}
                    for renamed_string in renameColumnString:
                        renamed_string = renamed_string.split(' as ')
                        renamed_cols.update({renamed_string[0]: renamed_string[1]})
                    # Rename columns using a loop or list comprehension
                    line_item = ''
                    for old_name, new_name in renamed_cols.items():
                        line_item += f"""
{df} = {df}.withColumnRenamed('{old_name}', '{new_name}')
{df}.show()
"""
                    return line_item, True
                else:
                    return "Key Value pair errors in rename column errors", False
            else:
                return "Key Value pair errors in rename column errors", False
        except Exception as e:
            return str(e), False
        
    def selectLeft(**kwargs):
        try:
            requiredKeys = ['df', 'oldColumn', 'newColumn', 'endIndex']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    df, oldColumn, newColumn, endIndex = values
                    lines = f"""
{df} = {df}.withColumn('{newColumn}', substring(col('{oldColumn}'), 0, {endIndex}))
{df}.show()
"""
                    return lines, True
                else:
                    return "Key Value pair errors select left function", False
            else:
                return "Key Value pair errors select left function", False
        
        except Exception as e:
            return str(e), False
    
    def selectRight(**kwargs):
        try:
            requiredKeys = ['df', 'oldColumn', 'newColumn', 'startIndex']
            if all(key in kwargs.keys() for key in requiredKeys):
                values = [kwargs.get(key) for key in requiredKeys]
                if all(value != '' for value in values):
                    df, oldColumn, newColumn, startIndex = values
                    lines = f"""
{df} = {df}.withColumn('{newColumn}', substring({df}['{oldColumn}'], length({df}['{oldColumn}']) - {startIndex}, length({df}['{oldColumn}'])))
{df}.show()
"""
                    return lines, True
                else:
                    return "Key Value pair errors select right function", False
            else:
                return "Key Value pair errors select right function", False
        
        except Exception as e:
            return str(e), False

    def columnTrim(**kwargs):
        try:
            df_found = False
            oldColumn_found = False
            newColumn_found = False

            for key, value in kwargs.items():
                if key == 'df':
                    df_found = True
                    df = value

                if key == 'oldColumn':
                    oldColumn_found = True
                    oldColumn = value

                if key == 'newColumn':
                    newColumn_found = True
                    newColumn = value
            
            if df_found & oldColumn_found & newColumn_found:
                lines = f"""
{df} = {df}.withColumn('{newColumn}', trim(regexp_replace(col("{oldColumn}"), r"\s+", " ")))
{df}.show()
"""
                return lines, True
            else:
                return "Key Value pair errorsin column trim function", False
        
        except Exception as e:
            return str(e), False

    def getLength(**kwargs):
        try:
            df_found = False
            newColumn_found = False
            oldColumn_found = False

            for key, value in kwargs.items():
                if key == 'df':
                    df_found = True 
                    df = value
                
                if key == 'oldColumn':
                    oldColumn_found = True
                    oldColumn = value
                
                if key == 'newColumn':
                    newColumn_found = True
                    newColumn = value
            
            if newColumn_found & oldColumn_found & df_found:
                lines = f"""
{df} = {df}.withColumn('{newColumn}', length('{oldColumn}'))
{df}.show()
"""
                return lines, True
            else:
                return "Key Value pair errors in getLength", False
        
        except Exception as e:
            return str(e), False

    def dropDuplicates(**kwargs):
        try:
            oldDf_found =  False
            newDf_found = False
            columnString_found = False

            for key,value in kwargs.items():
                if key ==  'newDf':
                    newDf_found = True
                    newDf = value

                if key == 'oldDf':
                    oldDf_found = True
                    oldDf = value
                
                if key == 'columnString':
                    columnString_found = True
                    columnString = value
            
            if columnString_found & oldDf_found & newDf_found:
                columnString = columnString.split(',')
                columnString_list = list(map(lambda x: x.strip(), columnString))

                lines = f"""
{newDf} = {oldDf}.dropDuplicates({columnString_list})
{newDf}.show()
"""
                return lines, True
            else: 
                return "Key Value pair errors in drop duplicates", False
        
        except Exception as e:
            return str(e), False
        
    def exactVLookup(**kwargs):
        try:
            requiredKeys = ['resultDf', 'leftDf', 'rightDf', 'column', 'columnsFromRight']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                resultDf, leftDf, rightDf, column, columnsFromRight = [kwargs.get(key) for key in requiredKeys]

                column_list = column.split(',')
                for i in range(len(column_list)):
                    column_list[i] = column_list[i].strip()

                columnsFromRightDf_list = columnsFromRight.split(',')
                for i in range(len(columnsFromRightDf_list)):
                    columnsFromRightDf_list[i] = columnsFromRightDf_list[i].strip()

                merged_column_list = column_list + columnsFromRightDf_list
                    
                lines = f"""
selected_df = {rightDf}.select(*{merged_column_list})
{resultDf} = {leftDf}.join(selected_df, on = {column_list}, how = 'left_outer')
{resultDf}.show()
"""
                return lines, True
            else:
                return "Key Value pair errors in exact Vlookup", False
        except Exception as e:
            return str(e), False
        
    def createNewColumnWithDefaultValue(**kwargs):
        try:
            df_found = False
            defaultValue_found = False
            columnName_found = False

            for key,value in kwargs.items():
                if key == 'df':
                    df_found = True
                    df = value
                if key == 'defaultValue':
                    defaultValue_found = True
                    defaultValue = value
                if key == 'columnName':
                    columnName_found = True
                    columnName = value

            if df_found & defaultValue_found & columnName_found:
                lines = f"""
{df} = {df}.withColumn('{columnName}', lit("{defaultValue}"))
{df}.show()
"""
                return lines, True
            else:
                return "key value pair error in createNewColumnWithDefaultValue function", False
        except Exception as e:
            return str(e), False

    def copyColToNewCol(**kwargs):
        try:
            df_found = False
            newColumn_found = False
            oldColumn_found = False

            for key, value in kwargs.items():
                if key == 'df':
                    df_found = True 
                    df = value
                
                if key == 'oldColumn':
                    oldColumn_found = True
                    oldColumn = value
                
                if key == 'newColumn':
                    newColumn_found = True
                    newColumn = value
            
            if newColumn_found & oldColumn_found & df_found:
                lines = f"""
{df} = {df}.withColumn('{newColumn}', col('{oldColumn}'))
{df}.show()
"""
                return lines, True
            else:
                return "Key Value pair errors in copy to new column function", False
        except Exception as e:
            return str(e), False
        
    def dropnaSubset(**kwargs):
        try:
            requiredKeys = ['df', 'columnString']
            if all(key in kwargs.keys() for key in requiredKeys):
                df = kwargs['df']
                columnString = kwargs['columnString']

                columnString = columnString.split(',')
                column_list = list(map(lambda x: x.strip(), columnString))
                lines =  f"""
################### dropna with subset #######################
{df} = {df}.dropna(subset={column_list})
{df}.show()
"""
                return lines, True
            else:
                return "Key Value pair error in dropna Subset Function", False
        except Exception as e:
            return str(e), False
        
    def leadColumn(**kwargs):
        try:
            requiredKeys = ['df','oldColumn', 'newColumn', 'leadBy']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df = kwargs['df']
                oldColumm = kwargs['oldColumm']
                newColumn = kwargs['newColumn']
                leadBy = int(kwargs['leadBy'])

                lines = f"""
windowSpec = Window.orderBy('{oldColumm}')
{df} = {df}.withColumn('{newColumn}', lead('{oldColumm}',{leadBy}).over(windowSpec))
{df} = {df}.withColumn('{newColumn}', when(col('{newColumn}').isNull(), lit(np.inf)).otherwise(col('{newColumn}')))
{df}.show()
"""
                return lines, True
            else:
                return "Key Value pair error in leadColumn Function", False
        except Exception as e:
            return str(e), False

    def lagColumn(**kwargs):
        try:
            requiredKeys = ['df','oldColumn', 'newColumn', 'lagBy']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df = kwargs['df']
                oldColumm = kwargs['oldColumm']
                newColumn = kwargs['newColumn']
                lagBy = int(kwargs['lagBy'])

                lines = f"""
windowSpec = Window.orderBy('{oldColumm}')
{df} = {df}.withColumn('{newColumn}', lag('{oldColumm}',{lagBy}).over(windowSpec))
{df} = {df}.withColumn('{newColumn}', when(col('{newColumn}').isNull(), lit(np.inf)).otherwise(col('{newColumn}')))
{df}.show()
"""
                return lines, True
            else:
                return "Key Value pair error in leadColumn Function", False
        except Exception as e:
            return str(e), False
        
    def dateDiff(**kwargs):
        try:
            requiredKeys = ['resultDf', 'df', 'newColumn', 'oldColumn', 'lastDay']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                resultDf, df, newColumn, oldColumn, lastDay = [kwargs.get(key) for key in requiredKeys]
                if lastDay:
                    lines = f"""
{resultDf} = {df}.withColumn("{newColumn}", datediff(last_day(current_date()), col("{oldColumn}")))
{resultDf}.show()
"""
                else:
                    lines = f"""
{resultDf} = {df}.withColumn("{newColumn}", datediff(current_date(), col("{oldColumn}")))
{resultDf}.show()
"""
                lines = f"""
wanted_date = datetime(2024,2,29)
{resultDf} = {df}.withColumn("{newColumn}", datediff(lit(wanted_date), col("{oldColumn}")))
{resultDf}.show()
"""
                return lines, True
            else:
                return "Key Value pair error in dateDiff Function", False
        except Exception as e:
            return str(e), False

    def approxVlookup(**kwargs):
        try:
            requiredKeys = ['resultDf', 'leftDf', 'rightDf', 'leftDfCol', 'rightDfCol', 'matchPercentage']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                resultDf, leftDf, rightDf, leftDfCol, rightDfCol, matchPercentage = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
def similarity(s1, s2):
    return process.extractOne(s1, [s2])[1]

# Register the function as a UDF
similarity_udf = spark.udf.register("similarity_udf", similarity, IntegerType())

# Calculate the similarity score
{resultDf} = {leftDf}.crossJoin({rightDf}) \
                .withColumn('similarity_score', similarity_udf(col('{leftDfCol}'), col('{rightDfCol}')))
{resultDf}.show()
# Filter the result based on the similarity score threshold
{resultDf} = {resultDf}.filter(col('similarity_score') >= {int(matchPercentage)})
{resultDf}.show()
"""
                return lines, True
            else:
                return "Error in key Value Pairs in approx vlookup", False
        except Exception as e:
            return str(e), False
        
    def writeExcel(**kwargs):
        try:
            requiredKeys = ['fileName', 'dfSheetNameString', 'temp_dir']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                fileName, dfSheetNameString, temp_dir = [kwargs.get(key) for key in requiredKeys]
                df_list = []
                sheet_name_list = []
                dfAndSheetName = dfSheetNameString.split(', ')
                for pair in dfAndSheetName:
                    df, sheet = pair.split("|")
                    df_list.append(df)
                    sheet_name_list.append(sheet)

                df_list = '[' + ', '.join(df_list) + ']'
                # create_directory_if_not_exists(f'{temp_dir}/outputs')
                fileName = f'{temp_dir}/output/{fileName}'

                lines = f'''
save_dataframes_to_excel({df_list}, {sheet_name_list}, "{fileName}")
'''
                return lines, True
            else:
                return "Error in key Value Pairs in writeExcel function", False
        except Exception as e:
            return str(e), False
        
    def writeExcelPreview(**kwargs):
        try:
            requiredKeys = ['fileName', 'dfSheetNameString', 'temp_dir']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                fileName, dfSheetNameString, temp_dir = [kwargs.get(key) for key in requiredKeys]
                df_list = []
                sheet_name_list = []
                dfAndSheetName = dfSheetNameString.split(', ')
                for pair in dfAndSheetName:
                    df, sheet = pair.split("|")
                    df_list.append(df)
                    sheet_name_list.append(sheet)

                df_list = '[' + ', '.join(df_list) + ']'
                fileName_joint = f'{temp_dir}/{fileName}'

                lines = f'''
save_dataframes_to_excel({df_list}, {sheet_name_list}, "{fileName_joint}")
'''
                return lines, fileName, True
            else:
                return "Error in key Value Pairs in writeExcel function", False, False
        except Exception as e:
            return str(e), False

    def writeCsv(**kwargs):
        pass

    def if_then_else(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'newColumn', 'otherwiseValue','conditions', 'isNumberOtherwiseValue']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newDf, newColumn, otherwiseValue, conditions, isNumberOtherwiseValue = [kwargs.get(key) for key in requiredKeys]
                condition_list = []
                for condition in conditions:
                    if 'group' in condition:
                        sub_str = ""
                        for sub_cond in condition['group']:
                            if not(sub_cond['isNumber']):
                                value = f"'{sub_cond['value']}'"
                            else:
                                value = sub_cond['value']

                            if sub_cond['logicalOperator'] == 'AND':
                                logicalOperator = '&'
                            elif sub_cond['logicalOperator'] == 'OR':
                                logicalOperator = '|'
                            else:
                                logicalOperator = ''
                            if sub_cond['columnOperation'] != '':
                                sub_str += f"({sub_cond['columnOperation']}(col('{sub_cond['columnName']}')) {sub_cond['operator']} {value}) {logicalOperator} "
                            else:
                                if sub_cond['operator'] == 'contains':
                                    sub_str += f"col('{sub_cond['columnName']}').contains({value}) {logicalOperator}"
                                else:
                                    sub_str += f"(col('{sub_cond['columnName']}') {sub_cond['operator']} {value}) {logicalOperator}"
                        if not(condition['isNumberOutput']):
                            output = f", '{condition['newValue']}'"
                        else:
                            output = f", {condition['newValue']}"

                        sub_str = sub_str.strip() + output
                        condition_list.append(sub_str)
                    else:
                        if not(condition['isNumber']):
                            value = f"'{condition['value']}'"
                        else:
                            value = condition['value']

                        if not(condition['isNumberOutput']):
                            output = f", '{condition['newValue']}'"
                        else:
                            output = f", {condition['newValue']}"

                        if condition['columnOperation'] != "":
                            strr = f"{condition['columnOperation']}(col('{condition['columnName']}')) {condition['operator']}"
                        else:
                            if condition['operator'] == 'contains':
                                strr = f"col('{condition['columnName']}').contains({value})"
                            else:
                                strr = f"col('{condition['columnName']}') {condition['operator']} {value}"

                        sub_str = strr.strip() + output
                        condition_list.append(sub_str)
                if isNumberOtherwiseValue:
                    otherwiseValue = f".otherwise({otherwiseValue})"
                else:
                    otherwiseValue = f".otherwise('{otherwiseValue}')"
                when_statement = ""
                for cond in condition_list:
                    when_statement += f".when({cond})"
                when_statement = when_statement[1:] + otherwiseValue
                lines = f'''
{newDf} = {oldDf}.withColumn('{newColumn}', {when_statement})
{newDf}.show()
'''
                return lines, True
            else:
                return "Kwargs error in if-then-else", False
        except Exception as e:
            return str(e), False
    
    def ageing(**kwargs):
        try:
            requiredKeys = ['masterDf', 'ageingReferenceColumn','agingDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                masterDf, agingReferenceColumn, agingDf = [kwargs.get(key) for key in requiredKeys]
                lines = f'''
windowSpec = Window.orderBy('Range')
{agingDf} = {agingDf}.withColumn('Range+1', lead('Range',1).over(windowSpec))
{agingDf} = {agingDf}.withColumn('Range+1', when(col('Range+1').isNull(), lit(np.inf)).otherwise(col('Range+1')))
{agingDf} = {agingDf}.select('Range','Range+1','Aging')
{agingDf}.show()
{masterDf} = {masterDf}.join({agingDf}, 
                    ({masterDf}['{agingReferenceColumn}'] >= {agingDf}['Range']) &
                    ({masterDf}['{agingReferenceColumn}'] < {agingDf}['Range+1']),
                    "left_outer")
{masterDf} = {masterDf}.drop("Range", "Range+1")
{masterDf}.show()
'''
                return lines, True
            else:
                return "key value pair error in aging", False
        except Exception as e:
            return str(e), False
        
    def groupBy(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'groupByColumnString', 'aggregateFunctions', 'verticalSummation']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newDf, groupByColumnString, aggregateFunctions, verticalSummation = [kwargs.get(key) for key in requiredKeys]
                agg_string = ''
                for function in aggregateFunctions:
                    if function['value'] != '':
                        sub_string = f" {function['aggregate']}('{function['key']}').alias('{function['value']}'),"
                    else:
                        sub_string = f" {function['aggregate']}('{function['key']}'),"
                    agg_string += sub_string
                agg_string = agg_string[:-1].strip()
                groupByColumnString = groupByColumnString.split(',')
                columnString = ""
                for column in groupByColumnString:
                    columnString += f" '{column.strip()}',"

                columnString = columnString[:-1].strip()
                lines = f'''
{newDf} = {oldDf}.groupBy({columnString}).agg({agg_string.strip()})
{newDf}.show()
'''
                '''
                if verticalSummation:

                    lines += f"""
# Identify numeric columns
numeric_cols = [field.name for field in {newDf}.schema.fields if isinstance(field.dataType, DoubleType)]

# Calculate the sum for numeric columns
sum_df = {newDf}.select([sum(col(c)).alias(c) for c in numeric_cols])

# Collect the sum values to create a new row
sum_values = sum_df.collect()[0].asDict()

# Create a new row with "Grand Total" and the sum values
grand_total_row = {{col: "Grand Total" if col not in numeric_cols else sum_values[col] for col in {newDf}.columns}}

# Convert the grand total row to a DataFrame
grand_total_df = spark.createDataFrame([grand_total_row])

# Union the original DataFrame with the grand total DataFrame
{newDf} = {newDf}.union(grand_total_df)

# Show the final DataFrame
{newDf}.show()
"""
'''
                return lines, True
            else:
                return "key value pair error in pivot", False
        except Exception as e:
            return str(e), False
        
    def pivot(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'groupByColumnString', 'pivotColumn', 'aggregateFunctions', 'horizontalSummation', 'verticalSummation']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newDf, groupByColumnString, pivotColumn, aggregateFunctions, horizontalSummation, verticalSummation = [kwargs.get(key) for key in requiredKeys]
                agg_string = ''
                for function in aggregateFunctions:
                    if function['value'] != '':
                        sub_string = f" {function['aggregate']}('{function['key']}').alias('{function['value']}'),"
                    else:
                        sub_string = f" {function['aggregate']}('{function['key']}'),"
                    agg_string += sub_string
                agg_string = agg_string[:-1].strip()
                groupByColumnString = groupByColumnString.split(',')
                columnString = ""
                for column in groupByColumnString:
                    columnString += f" '{column.strip()}',"

                columnString = columnString[:-1].strip()
                lines = f'''
{newDf} = {oldDf}.groupBy({columnString}).pivot('{pivotColumn}').agg({agg_string.strip()})
'''
                if verticalSummation:
                    lines += f''' 
agg_cols = {newDf}.columns[1:]
first_col = {newDf}.columns[0]
rollup_df = {newDf}.rollup().sum()
rollup_df.show()

renamed_df = reduce(
    lambda rollup_df, idx: rollup_df.withColumnRenamed(rollup_df.columns[idx], agg_cols[idx]), 
    range(len(rollup_df.columns)), rollup_df
)
renamed_df.show()

renamed_df = renamed_df.withColumn(first_col, lit('Grand Total'))
renamed_df.show()

{newDf} = {newDf}.unionByName(renamed_df)
{newDf}.show()
'''

                if horizontalSummation:
                    lines += f'''
agg_cols = {newDf}.columns[1:]
first_col = {newDf}.columns[0]
{newDf} = {newDf}.na.fill(value = 0).withColumn('Grand Total', reduce(add, [col(c) for c in agg_cols]))
{newDf} = {newDf}.withColumnRenamed(first_col, 'Row Labels')
{newDf}.show()
'''
                return lines, True
            else:
                return "key value pair error in pivot", False
        except Exception as e:
            return str(e), False
        
    def getSumOfColumn(**kwargs):
        try:
            requiredKeys = ['df', 'columnName', 'resultDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, columnName, resultDf = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
{resultDf} = {df}.agg(
    sum(col('{columnName}')).alias("sum"))

{resultDf}.show()
"""
                return lines, True
            else:
                return "key value pair error in sum of column", False
        except Exception as e:
            return str(e), False
        
    def getAvgOfColumn(**kwargs):
        try:
            requiredKeys = ['df', 'columnName', 'resultDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, columnName, resultDf = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
{resultDf} = {df}.agg(
    avg(col('{columnName}')).alias("average"))

{resultDf}.show()
"""
                return lines, True
            else:
                return "key value pair error in avg of column", False
        except Exception as e:
            return str(e), False
        
    def getMinOfColumn(**kwargs):
        try:
            requiredKeys = ['df', 'columnName', 'resultDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, columnName, resultDf = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
{resultDf} = {df}.agg(
    min(col('{columnName}')).alias("min"))

{resultDf}.show()
"""
                return lines, True
            else:
                return "key value pair error in min of column", False
        except Exception as e:
            return str(e), False
        
    def getMaxOfColumn(**kwargs):
        try:
            requiredKeys = ['df', 'columnName', 'resultDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, columnName, resultDf = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
{resultDf} = {df}.agg(
    max(col('{columnName}')).alias("max"))

{resultDf}.show()
"""
                return lines, True
            else:
                return "key value pair error in max of column", False
        except Exception as e:
            return str(e), False
        
    def getCountOfColumn(**kwargs):
        try:
            requiredKeys = ['df', 'columnName', 'resultDf']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, columnName, resultDf = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
{resultDf} = {df}.agg(
    count(col('{columnName}')).alias("count"))

{resultDf}.show()
"""
                return lines, True
            else:
                return "key value pair error in count of column", False
        except Exception as e:
            return str(e), False
        
    def startIf(**kwargs):
        try:
            requiredKeys = ['code']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                code = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
{code[0]}
"""
                return lines, True
            else:
                return "key value pair error start if condition", False
        except Exception as e:
            return str(e), False
        
    def elIf(**kwargs):
        try:
            requiredKeys = ['code']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                code = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
{code[0]}
"""
                return lines, True
            else:
                return "key value pair error start if condition", False
        except Exception as e:
            return str(e), False
        
    def elSe():
        lines = f"""
else:
"""
        return lines, True

    def addColumns(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newColumn', 'columnString']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newColumn, columnString = [kwargs.get(key) for key in requiredKeys]
                columnsToAdd = columnString.split(',')
                for i in range(len(columnsToAdd)):
                    columnsToAdd[i] = columnsToAdd[i].strip()
                columnstr = ''
                for column in columnsToAdd:
                    columnstr += f'col("{column}")+'

                columnstr = columnstr[:-1]
                lines = f'''
{oldDf} = {oldDf}.withColumn("{newColumn}", {columnstr})
'''
                return lines, True
            else:
                return "key value pair error start if condition", False
        except Exception as e:
            return str(e), False
        
    def subtractColumns(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newColumn', 'columnString']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newColumn, columnString = [kwargs.get(key) for key in requiredKeys]
                columnsToSubtract = columnString.split(',')
                for i in range(len(columnsToSubtract)):
                    columnsToSubtract[i] = columnsToSubtract[i].strip()
                columnstr = ''
                for column in columnsToSubtract:
                    columnstr += f'col("{column}")-'

                columnstr = columnstr[:-1]
                lines = f'''
{oldDf} = {oldDf}.withColumn("{newColumn}", {columnstr})
{oldDf}.show()
'''
                return lines, True
            else:
                return "key value pair error start if condition", False
        except Exception as e:
            return str(e), False
        
    def multiplyColumns(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newColumn', 'columnString']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newColumn, columnString = [kwargs.get(key) for key in requiredKeys]
                columnsToMultiply = columnString.split(',')
                for i in range(len(columnsToMultiply)):
                    columnsToMultiply[i] = columnsToMultiply[i].strip()
                columnstr = ''
                for column in columnsToMultiply:
                    columnstr += f'col("{column}")*'

                columnstr = columnstr[:-1]
                lines = f'''
{oldDf} = {oldDf}.withColumn("{newColumn}", {columnstr})
{oldDf}.show()
'''
                return lines, True
            else:
                return "key value pair error start if condition", False
        except Exception as e:
            return str(e), False
        
    def divideColumns(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newColumn', 'columnString']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newColumn, columnString = [kwargs.get(key) for key in requiredKeys]
                columnsToDivide = columnString.split(',')
                for i in range(len(columnsToDivide)):
                    columnsToDivide[i] = columnsToDivide[i].strip()
                columnstr = ''
                for column in columnsToDivide:
                    columnstr += f'col("{column}")/'

                columnstr = columnstr[:-1]
                lines = f'''
{oldDf} = {oldDf}.withColumn("{newColumn}", {columnstr})
{oldDf}.show()
'''
                return lines, True
            else:
                return "key value pair error start if condition", False
        except Exception as e:
            return str(e), False
        
    def findAndReplaceInColumn(**kwargs):
        try:
            requiredKeys = ['oldDf', 'newDf', 'configList']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newDf, findandreplacelist = [kwargs.get(key) for key in requiredKeys]
                
                lines = f'{newDf} = {oldDf}'
                for entry in findandreplacelist:
                    lines += f""".withColumn("{entry['newColumn']}", regexp_replace("{entry['oldColumn']}","{entry['oldValue']}","{entry['newValue']}"))
{newDf}.show()"""
                    pass
                return lines, True
            else:
                return "key value pair error findand replace column", False
        except Exception as e:
            return str(e), False
    
    def sortDf(**kwargs):
        '''
        Takes 3 arguments for
        1. old Table Name. --> str
        2. new Table name if you have to assing to. --> str
        3. column dict list --> list of dict
            [
                {'columnName': <columnName>,
                 'sortType': <desc/asc>},
                {'columnName': <columnName>,
                 'sortType': <desc/asc>}
            ]
        '''
        try:
            requiredKeys = ['oldDf', 'newDf', 'columnDictList']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                oldDf, newDf, columnDictList = [kwargs.get(key) for key in requiredKeys]
                # prepare sorting string
                sorting_str = ''
                for entry in columnDictList:
                    column_name = entry['columnName']
                    sort_type = entry['sortType']
                    sorting_str += f' {sort_type}("{column_name}"),'

                sorting_str = sorting_str.strip()[:-1]

                lines = f"""
{newDf} = {oldDf}.orderBy({sorting_str})
{newDf}.show()
"""
                return lines, True
            else:
                return "key value pair error start if condition", False
        except Exception as e:
            return str(e), False
        
    def lastBlock():
        lines = """
    completion_status = True
    completion_output = "Success"
except Exception as e:
    completion_status = False
    completion_output = str(e)
    print(str(e))
finally:
    transformation_job_id = os.getenv("TRANSFORMATION_JOB_ID")
    print(transformation_job_id)
    payload = json.dumps({
        "transformationJobId": transformation_job_id,
        "databricksJobStatus": completion_status,
        "databricksOutput": completion_output
    })
    headers = {
        'Content-Type': 'application/json'
    }
    url = os.getenv("APPLICATION_CALLBACK_URL")

    response = requests.request(method = "POST", url = url, headers=headers, data=payload)
    print(payload)
    print(response.text)
"""
        return lines, True
    
    def combineTables(**kwargs):
        try:
            requiredKeys = ['resultDf', 'inputDataframeList']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                resultDf, inputDataframeList = [kwargs.get(key) for key in requiredKeys]
                """
                Merges multiple DataFrames one below the other.
                
                Parameters:
                dataframes (List[DataFrame]): List of DataFrames to be merged.

                Returns:
                DataFrame: A single DataFrame that is the result of merging all input DataFrames.
                """
                inputDataframeList = '[' + ','.join(inputDataframeList) + ']'
                lines = f"""
list_of_df = {inputDataframeList}
if not {inputDataframeList}:
    raise ValueError("The list of DataFrames is empty")

# Start with the first DataFrame
{resultDf} = list_of_df[0]

# Iterate through the remaining DataFrames and merge them
for df in list_of_df[1:]:
    {resultDf} = {resultDf}.unionByName(df, allowMissingColumns=True)

{resultDf}.show()
"""
                return lines, True
            else:
                return f"error in combineTables: {str(e)}"
        except Exception as e:
            return str(e), False

    def writeToADLSGen2(**kwargs):
        try:
            requiredKeys = ['tableDict', 'hiveSchemaName']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                tableDict, hiveSchemaName = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
spark.sql("CREATE NAMESPACE IF NOT EXISTS {os.getenv("AZURE_CATALOG")}.{hiveSchemaName}")
"""
                for key, value in tableDict:
                    lines += f"""
{key}.write.mode("overwrite").saveAsTable(name = '{os.getenv("AZURE_CATALOG")}.{hiveSchemaName}.{value}', format='parquet')
"""
                return lines, True
            else: 
                return "error in generating code", False
        except Exception as e:
            return str(e), False
        
    def zipFiles(temp_dir, config_list, sftp_details, n1, recon_name):
        try:
            lines = f"""
temp_dir = '{temp_dir}'
config_list = {config_list}
date = '{n1}'
recon_name = '{recon_name}'
file_list = []
for config in config_list:
    contains = None
    startsWith=None
    endsWith=None
    equals=None
    directory = temp_dir
    if config['searchCriteria'] == "contains":
        contains = config['value']
    elif config['searchCriteria'] == "startsWith":
        startsWith = config['value']
    elif config['searchCriteria'] == "endsWith":
        endsWith = config['value']
    else:
        equals = config['value']

    file_list.append(getFilePathOnADLS(directory=directory,
                                    contains=contains,
                                    startsWith=startsWith,
                                    endsWith=endsWith,
                                    equals=equals))
file_list = [x for x in file_list if x is not None]

adls_zip_file_path = f"{{temp_dir}}/outputzips/{{recon_name}}/{{date}}_output.zip"

zipFilesOnADLS(file_paths=file_list,zip_file_path = adls_zip_file_path)

sftp_details['key_file'] = os.getenv("SFTP_KEY_FILE")
sftp_details['password'] = os.getenv("SFTP_PASSWORD")
if sftp_details['key_file']:
    sftp_details['password'] = None
else:
    sftp_details['key_file'] = None

sftp, transport = get_sftp(hostname=sftp_details['hostname'],
                username=sftp_details['username'],
                password=sftp_details['password'],
                port=int(sftp_details['port']),
                key_file=sftp_details['key_file'])
sftp_mkdirs(sftp=sftp, remote_path=sftp_details['outputfileLocation'])

blob_service_client = BlobServiceClient(account_url=f"https://{{os.getenv('AZURE_STORAGE_ACCOUNT_NAME')}}.blob.core.windows.net", credential=os.getenv("AZURE_STORAGE_ACCOUNT_KEY"))
container_client = blob_service_client.get_container_client(os.getenv("AZURE_BLOB_CONTAINER_NAME"))
blob_client = container_client.get_blob_client(adls_zip_file_path)

# Stream blob data directly into memory
blob_data = blob_client.download_blob().readall()

# Write to SFTP using in-memory buffer
with BytesIO(blob_data) as file_stream:
    sftp.putfo(file_stream, f"{{sftp_details['outputfileLocation']}}/{{date}}.zip")

# sftp.put(adls_zip_file_path, f"{{sftp_details['outputfileLocation']}}/{{date}}.zip")
# shutil.rmtree(periodic_run_zip)
"""
            return lines, True
        except Exception as e:
            return str(e), False
        
    def mid(**kwargs):
        try:
            requiredKeys = ['df', 'oldColumn', 'newColumn', 'startPosition', 'buffer']
            if all(keys in kwargs.keys() for keys in requiredKeys):
                df, oldColumn, newColumn,startPosition, buffer = [kwargs.get(key) for key in requiredKeys]
                lines = f"""
{df} = {df}.withColumn("{newColumn}", substring("{oldColumn}", {startPosition}, {buffer}))
{df}.show()
"""
                return lines, True
            else:
                return "Key-value pair error in mid function", False
        except Exception as e:
            return f"Error in PreviewRepoService mid() {str(e)}", False